#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
import xlrd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_HEADERS = ["meno", "priezvisko", "email", "telefón"]
EMAIL_RE = re.compile(r"^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$", re.IGNORECASE)
EMAIL_EXTRACT_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
HEADER_ALIASES = {
    "meno": [
        "meno",
        "first_name",
        "firstname",
        "first",
        "given_name",
        "givenname",
        "krstne_meno",
        "krstne",
    ],
    "priezvisko": [
        "priezvisko",
        "last_name",
        "lastname",
        "surname",
        "family_name",
        "familyname",
        "last",
    ],
    "cele_meno": [
        "cele_meno",
        "celemeno",
        "full_name",
        "fullname",
        "name",
        "contact_name",
        "meno_a_priezvisko",
    ],
    "email": [
        "email",
        "email_address",
        "e_mail",
        "e-mail",
        "mail",
    ],
    "telefon": [
        "telefon",
        "telefonne_cislo",
        "telefón",
        "phone",
        "phone_number",
        "mobile",
        "mobil",
        "telephone",
        "gsm",
        "tel",
    ],
}


@dataclass(frozen=True)
class InputRow:
    file_path: str
    row_number: int
    raw: dict[str, str]


@dataclass
class ContactRecord:
    source: InputRow
    meno: str
    priezvisko: str
    email: str
    telefon: str

    @property
    def full_name_key(self) -> str:
        return normalize_name_key(f"{self.meno} {self.priezvisko}".strip())

    def to_output_row(self) -> dict[str, str]:
        return {
            "meno": self.meno,
            "priezvisko": self.priezvisko,
            "email": self.email,
            "telefón": self.telefon,
        }

    def to_audit_row(self) -> dict[str, str]:
        return {
            "meno": self.meno,
            "priezvisko": self.priezvisko,
            "email": self.email,
            "telefón": self.telefon,
            "zdroj": self.source.file_path,
            "riadok": str(self.source.row_number),
        }


class UnionFind:
    def __init__(self, size: int) -> None:
        self.parent = list(range(size))
        self.rank = [0] * size

    def find(self, item: int) -> int:
        while self.parent[item] != item:
            self.parent[item] = self.parent[self.parent[item]]
            item = self.parent[item]
        return item

    def union(self, left: int, right: int) -> None:
        root_left = self.find(left)
        root_right = self.find(right)
        if root_left == root_right:
            return
        if self.rank[root_left] < self.rank[root_right]:
            self.parent[root_left] = root_right
            return
        if self.rank[root_left] > self.rank[root_right]:
            self.parent[root_right] = root_left
            return
        self.parent[root_right] = root_left
        self.rank[root_left] += 1


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def normalize_header(value: str) -> str:
    base = strip_accents(str(value or "").strip().lower())
    base = re.sub(r"[^a-z0-9]+", "_", base)
    return base.strip("_")


def normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def smart_name_case(value: str) -> str:
    value = normalize_whitespace(value)
    if not value:
        return ""

    def format_token(token: str) -> str:
        pieces = re.split(r"([-'`])", token)
        return "".join(
            piece.capitalize() if index % 2 == 0 else piece
            for index, piece in enumerate(pieces)
        )

    return " ".join(format_token(token) for token in value.split(" "))


def normalize_name(value: str) -> str:
    text = normalize_whitespace(value)
    if not text:
        return ""
    if EMAIL_EXTRACT_RE.search(text):
        return ""
    if normalize_phone(text):
        return ""
    if looks_like_date(text):
        return ""
    text = re.sub(r"\s+", " ", text)
    return smart_name_case(text)


def normalize_name_key(value: str) -> str:
    base = strip_accents(normalize_whitespace(value).lower())
    base = re.sub(r"[^a-z ]+", " ", base)
    return re.sub(r"\s+", " ", base).strip()


def looks_like_date(value: str) -> bool:
    text = normalize_whitespace(value)
    digits = re.sub(r"\D", "", text)
    return bool(
        ":" in text
        or re.search(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b", text)
        or re.search(r"\b\d{4}-\d{2}-\d{2}\b", text)
        or re.fullmatch(r"(19|20)\d{6,14}", digits)
    )


def extract_email(value: str) -> str:
    text = normalize_whitespace(value).strip(";,")
    if not text:
        return ""
    match = EMAIL_EXTRACT_RE.search(text)
    if not match:
        return ""
    email = match.group(0).lower()
    return email if EMAIL_RE.fullmatch(email) else ""


def normalize_phone(value: str, default_country_code: str = "+421") -> str:
    text = normalize_whitespace(value)
    if not text or looks_like_date(text):
        return ""
    if EMAIL_EXTRACT_RE.search(text):
        return ""

    text = re.sub(r"(?i)\b(ext|extension|klapka|kl\.?)\s*\d+\b", "", text).strip()
    if not text:
        return ""

    has_plus = text.startswith("+")
    digits = re.sub(r"\D", "", text)
    if len(digits) < 9 or len(digits) > 15:
        return ""

    if text.startswith("00"):
        return f"+{digits[2:]}"

    if has_plus:
        return f"+{digits}"

    if len(digits) == 10 and digits.startswith("0"):
        return f"{default_country_code}{digits[1:]}"

    if len(digits) == 9:
        return f"{default_country_code}{digits}"

    if len(digits) >= 11:
        return f"+{digits}"

    return ""


def split_full_name(value: str) -> tuple[str, str]:
    normalized = normalize_name(value)
    if not normalized:
        return "", ""
    parts = normalized.split(" ")
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def resolve_name_fields(
    first_name_raw: str,
    last_name_raw: str,
    full_name_raw: str,
) -> tuple[str, str]:
    split_first, split_last = split_full_name(full_name_raw)

    first_name = normalize_name(first_name_raw) or split_first
    last_name = normalize_name(last_name_raw) or split_last

    first_key = normalize_name_key(first_name)
    last_key = normalize_name_key(last_name)

    # Priezvisko nesmie byť len kópia mena.
    if first_key and last_key and first_key == last_key:
        if split_last and normalize_name_key(split_last) != first_key:
            last_name = split_last
        else:
            last_name = ""

    # Ak bolo meno prázdne a máme z full_name lepšie rozdelenie, doplníme ho.
    if not first_name and split_first:
        first_name = split_first
    if not last_name and split_last and normalize_name_key(split_last) != normalize_name_key(first_name):
        last_name = split_last

    return first_name, last_name


def read_table(path: Path) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".xls":
        return read_xls(path)
    raise ValueError(f"Nepodporovaný formát súboru: {path.name}")


def read_table_from_bytes(file_name: str, file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return read_csv_bytes(file_bytes, file_name)
    if suffix == ".xlsx":
        return read_xlsx_bytes(file_bytes)
    if suffix == ".xls":
        return read_xls_bytes(file_bytes)
    raise ValueError(f"Nepodporovaný formát súboru: {file_name}")


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    return read_csv_bytes(path.read_bytes(), path.name)


def read_csv_bytes(raw_bytes: bytes, file_name: str = "upload.csv") -> tuple[list[str], list[list[str]]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"CSV súbor sa nepodarilo dekódovať: {file_name}")

    rows = list(csv.reader(io.StringIO(text), delimiter=detect_csv_delimiter(text)))
    return normalize_table_rows(rows)


def detect_csv_delimiter(text: str) -> str:
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        if dialect and dialect.delimiter:
            return dialect.delimiter
    except csv.Error:
        pass

    lines = [line for line in sample.splitlines() if line.strip()][:20]
    if not lines:
        return ","

    candidates = [";", ",", "\t", "|"]
    best = ","
    best_score = -1
    for delimiter in candidates:
        counts = [line.count(delimiter) for line in lines]
        score = sum(counts) + sum(1 for count in counts if count > 0) * 2
        if score > best_score:
            best_score = score
            best = delimiter
    return best


def read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    return read_xlsx_bytes(path.read_bytes())


def read_xlsx_bytes(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([normalize_cell(cell) for cell in row])
    return normalize_table_rows(rows)


def read_xls(path: Path) -> tuple[list[str], list[list[str]]]:
    return read_xls_bytes(path.read_bytes())


def read_xls_bytes(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    rows = []
    for row_index in range(sheet.nrows):
        row = [normalize_cell(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
        rows.append(row)
    return normalize_table_rows(rows)


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_whitespace(str(value))


def normalize_table_rows(rows: Sequence[Sequence[str]]) -> tuple[list[str], list[list[str]]]:
    if not rows:
        raise ValueError("Súbor je prázdny.")

    first_row = list(rows[0])
    headers = [
        normalize_whitespace(cell) or f"column_{index + 1}"
        for index, cell in enumerate(first_row)
    ]

    body = []
    for row in rows[1:]:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        normalized = [normalize_whitespace(padded[index]) for index in range(len(headers))]
        if any(normalized):
            body.append(normalized)
    return headers, body


def score_column(header_key: str, values: Sequence[str], logical_field: str) -> int:
    score = 0
    aliases = HEADER_ALIASES[logical_field]
    if header_key in aliases:
        score += 100
    elif any(alias in header_key for alias in aliases):
        score += 30

    samples = [value for value in values if value][:30]
    if logical_field == "email":
        score += sum(10 if extract_email(value) else -2 for value in samples)
    elif logical_field == "telefon":
        score += sum(8 if normalize_phone(value) else -3 for value in samples)
    elif logical_field == "cele_meno":
        score += sum(5 if len(normalize_name(value).split()) >= 2 else -2 for value in samples)
    else:
        score += sum(3 if len(normalize_name(value).split()) == 1 else -1 for value in samples)
    return score


def rank_columns(
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    logical_field: str,
) -> list[tuple[int, int]]:
    normalized_headers = [normalize_header(header) for header in headers]
    columns = list(zip(*rows)) if rows else [tuple() for _ in headers]
    ranked: list[tuple[int, int]] = []

    for index, header_key in enumerate(normalized_headers):
        values = columns[index] if columns else []
        ranked.append((score_column(header_key, values, logical_field), index))

    ranked.sort(key=lambda item: (item[0], -item[1]), reverse=True)
    return ranked


def detect_mapping(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> dict[str, int | None]:
    result: dict[str, int | None] = {}
    thresholds = {
        "meno": 25,
        "priezvisko": 25,
        "cele_meno": 25,
        "email": 45,
        "telefon": 45,
    }

    rankings = {
        logical_field: rank_columns(headers, rows, logical_field)
        for logical_field in ("meno", "priezvisko", "cele_meno", "email", "telefon")
    }

    for logical_field in ("cele_meno", "email", "telefon"):
        best_score, best_index = rankings[logical_field][0] if rankings[logical_field] else (float("-inf"), None)
        result[logical_field] = best_index if best_score >= thresholds[logical_field] else None

    used_name_columns: set[int] = set()
    for logical_field in ("meno", "priezvisko"):
        chosen_index = None
        for score, index in rankings[logical_field]:
            if score < thresholds[logical_field]:
                break
            if index in used_name_columns:
                continue
            chosen_index = index
            break
        result[logical_field] = chosen_index
        if chosen_index is not None:
            used_name_columns.add(chosen_index)

    if result["meno"] is not None and result["priezvisko"] is not None and result["meno"] == result["priezvisko"]:
        result["priezvisko"] = None

    return result


def build_contacts(input_files: Sequence[Path], default_country_code: str) -> tuple[list[ContactRecord], int]:
    contacts: list[ContactRecord] = []
    imported_rows = 0

    for path in input_files:
        headers, rows = read_table(path)
        mapping = detect_mapping(headers, rows)

        for row_number, row in enumerate(rows, start=2):
            imported_rows += 1
            raw = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}

            full_name = row[mapping["cele_meno"]] if mapping["cele_meno"] is not None else ""
            first_name, last_name = resolve_name_fields(
                row[mapping["meno"]] if mapping["meno"] is not None else "",
                row[mapping["priezvisko"]] if mapping["priezvisko"] is not None else "",
                full_name,
            )
            email = extract_email(row[mapping["email"]]) if mapping["email"] is not None else ""
            phone = normalize_phone(
                row[mapping["telefon"]] if mapping["telefon"] is not None else "",
                default_country_code=default_country_code,
            )

            contacts.append(
                ContactRecord(
                    source=InputRow(file_path=path.as_posix(), row_number=row_number, raw=raw),
                    meno=first_name,
                    priezvisko=last_name,
                    email=email,
                    telefon=phone,
                )
            )

    return contacts, imported_rows


def build_contacts_from_uploads(
    uploads: Sequence[tuple[str, bytes]],
    default_country_code: str,
) -> tuple[list[ContactRecord], int, list[dict[str, int | str | dict[str, int | None]]]]:
    contacts: list[ContactRecord] = []
    imported_rows = 0
    dataset_reports: list[dict[str, int | str | dict[str, int | None]]] = []

    for file_name, file_bytes in uploads:
        headers, rows = read_table_from_bytes(file_name, file_bytes)
        mapping = detect_mapping(headers, rows)
        valid_rows_in_file = 0

        for row_number, row in enumerate(rows, start=2):
            imported_rows += 1
            raw = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}

            full_name = row[mapping["cele_meno"]] if mapping["cele_meno"] is not None else ""
            first_name, last_name = resolve_name_fields(
                row[mapping["meno"]] if mapping["meno"] is not None else "",
                row[mapping["priezvisko"]] if mapping["priezvisko"] is not None else "",
                full_name,
            )
            email = extract_email(row[mapping["email"]]) if mapping["email"] is not None else ""
            phone = normalize_phone(
                row[mapping["telefon"]] if mapping["telefon"] is not None else "",
                default_country_code=default_country_code,
            )

            contact = ContactRecord(
                source=InputRow(file_path=file_name, row_number=row_number, raw=raw),
                meno=first_name,
                priezvisko=last_name,
                email=email,
                telefon=phone,
            )
            contacts.append(contact)
            if is_usable_contact(contact):
                valid_rows_in_file += 1

        dataset_reports.append(
            {
                "file_name": file_name,
                "total_rows": len(rows),
                "valid_contacts": valid_rows_in_file,
                "mapping": mapping,
            }
        )

    return contacts, imported_rows, dataset_reports


def is_usable_contact(contact: ContactRecord) -> bool:
    return bool(contact.email or contact.telefon)


def are_name_only_duplicates(left: ContactRecord, right: ContactRecord) -> bool:
    if not left.full_name_key or left.full_name_key != right.full_name_key:
        return False
    if not left.meno or not left.priezvisko:
        return False
    if left.email and right.email and left.email != right.email:
        return False
    if left.telefon and right.telefon and left.telefon != right.telefon:
        return False

    # Meno samotné nestačí. Tento fallback použijeme len vtedy, keď jeden z kontaktov
    # má iba email a druhý iba telefón, bez akéhokoľvek konfliktu v údajoch.
    complementary_channels = bool(left.email) != bool(right.email) and bool(left.telefon) != bool(right.telefon)
    return complementary_channels


def deduplicate_contacts(
    contacts: Sequence[ContactRecord],
) -> tuple[list[ContactRecord], int, list[dict[str, object]]]:
    if not contacts:
        return [], 0, []

    union_find = UnionFind(len(contacts))
    email_index: dict[str, int] = {}
    phone_index: dict[str, int] = {}
    name_buckets: dict[str, list[int]] = {}

    for index, contact in enumerate(contacts):
        if contact.email:
            existing = email_index.get(contact.email)
            if existing is not None:
                union_find.union(existing, index)
            else:
                email_index[contact.email] = index

        if contact.telefon:
            existing = phone_index.get(contact.telefon)
            if existing is not None:
                union_find.union(existing, index)
            else:
                phone_index[contact.telefon] = index

        if contact.full_name_key:
            name_buckets.setdefault(contact.full_name_key, []).append(index)

    for bucket in name_buckets.values():
        if len(bucket) < 2:
            continue
        for position, left_index in enumerate(bucket):
            for right_index in bucket[position + 1:]:
                if are_name_only_duplicates(contacts[left_index], contacts[right_index]):
                    union_find.union(left_index, right_index)

    groups: dict[int, list[ContactRecord]] = {}
    for index, contact in enumerate(contacts):
        root = union_find.find(index)
        groups.setdefault(root, []).append(contact)

    winners = [pick_best_contact(group) for group in groups.values()]
    winners.sort(key=sort_key)
    removed_duplicates = len(contacts) - len(winners)
    removed_audit = build_removed_duplicates_audit(groups)
    return winners, removed_duplicates, removed_audit


def build_removed_duplicates_audit(
    groups: dict[int, list[ContactRecord]],
) -> list[dict[str, object]]:
    audit_rows: list[dict[str, object]] = []

    for group in groups.values():
        if len(group) < 2:
            continue

        winner = pick_best_contact(group)
        reason = detect_group_reason(group)

        for contact in group:
            if contact is winner:
                continue
            audit_rows.append(
                {
                    "reason": reason,
                    "kept": winner.to_audit_row(),
                    "removed": contact.to_audit_row(),
                }
            )

    audit_rows.sort(
        key=lambda item: (
            normalize_name_key(str(item["removed"].get("priezvisko", ""))),
            normalize_name_key(str(item["removed"].get("meno", ""))),
            str(item["removed"].get("email", "")),
            str(item["removed"].get("telefón", "")),
        )
    )
    return audit_rows


def detect_group_reason(group: Sequence[ContactRecord]) -> str:
    emails = {contact.email for contact in group if contact.email}
    phones = {contact.telefon for contact in group if contact.telefon}
    if len(emails) == 1 and emails:
        return "zhoda emailu"
    if len(phones) == 1 and phones:
        return "zhoda telefónu"
    return "veľmi podobné meno a nekonfliktné kontaktné údaje"


def contact_quality_score(contact: ContactRecord) -> tuple[int, int, int, int, int]:
    filled_fields = sum(1 for value in (contact.meno, contact.priezvisko, contact.email, contact.telefon) if value)
    name_strength = int(bool(contact.meno)) + int(bool(contact.priezvisko))
    email_strength = 1 if contact.email else 0
    phone_strength = 1 if contact.telefon else 0
    raw_length = sum(len(value) for value in (contact.meno, contact.priezvisko, contact.email, contact.telefon))
    return filled_fields, name_strength, email_strength, phone_strength, raw_length


def pick_best_contact(group: Sequence[ContactRecord]) -> ContactRecord:
    return max(
        group,
        key=lambda contact: (
            contact_quality_score(contact),
            -group.index(contact),
        ),
    )


def sort_key(contact: ContactRecord) -> tuple[str, str, str, str]:
    return (
        normalize_name_key(contact.priezvisko),
        normalize_name_key(contact.meno),
        contact.email,
        contact.telefon,
    )


def export_csv(path: Path, contacts: Sequence[ContactRecord]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        for contact in contacts:
            writer.writerow(contact.to_output_row())


def export_xlsx(path: Path, contacts: Sequence[ContactRecord]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Kontakty"

    sheet.append(OUTPUT_HEADERS)
    for contact in contacts:
        sheet.append([contact.to_output_row()[header] for header in OUTPUT_HEADERS])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)

    workbook.save(path)


def write_report(path: Path, report: dict[str, int | str]) -> None:
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def process_uploads(
    uploads: Sequence[tuple[str, bytes]],
    default_country_code: str = "+421",
) -> dict[str, object]:
    all_contacts, imported_rows, dataset_reports = build_contacts_from_uploads(
        uploads,
        default_country_code=default_country_code,
    )
    usable_contacts = [contact for contact in all_contacts if is_usable_contact(contact)]
    dropped_without_channel = len(all_contacts) - len(usable_contacts)
    deduplicated_contacts, removed_duplicates, removed_audit = deduplicate_contacts(usable_contacts)

    return {
        "datasets": dataset_reports,
        "rows": [contact.to_output_row() for contact in deduplicated_contacts],
        "removed_duplicates": removed_audit,
        "report": {
            "pocet_importovanych_zaznamov": imported_rows,
            "pocet_odstranenych_duplicit": removed_duplicates,
            "pocet_vyradenych_bez_emailu_a_telefonu": dropped_without_channel,
            "pocet_finalnych_kontaktov": len(deduplicated_contacts),
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import, vyčistenie a deduplikácia kontaktov do jedného finálneho exportu."
    )
    parser.add_argument(
        "inputs",
        nargs="+",
        help="Cesty k vstupným CSV/XLSX/XLS súborom.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="output/final_contacts.xlsx",
        help="Cesta k finálnemu výstupu (.csv alebo .xlsx). Predvolené: output/final_contacts.xlsx",
    )
    parser.add_argument(
        "--report",
        default="output/final_contacts_report.json",
        help="Cesta k JSON reportu. Predvolené: output/final_contacts_report.json",
    )
    parser.add_argument(
        "--default-country-code",
        default="+421",
        help="Predvolená krajina pre lokálne telefónne čísla. Predvolené: +421",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_files = [Path(item).expanduser().resolve() for item in args.inputs]
    output_path = Path(args.output).expanduser().resolve()
    report_path = Path(args.report).expanduser().resolve()

    for path in input_files:
        if not path.exists():
            raise FileNotFoundError(f"Vstupný súbor neexistuje: {path}")

    all_contacts, imported_rows = build_contacts(input_files, args.default_country_code)
    usable_contacts = [contact for contact in all_contacts if is_usable_contact(contact)]
    dropped_without_channel = len(all_contacts) - len(usable_contacts)
    deduplicated_contacts, removed_duplicates, _removed_audit = deduplicate_contacts(usable_contacts)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.suffix.lower() == ".csv":
        export_csv(output_path, deduplicated_contacts)
    elif output_path.suffix.lower() == ".xlsx":
        export_xlsx(output_path, deduplicated_contacts)
    else:
        raise ValueError("Výstup musí byť vo formáte .csv alebo .xlsx")

    report = {
        "vstupne_subory": [path.as_posix() for path in input_files],
        "vystupny_subor": output_path.as_posix(),
        "pocet_importovanych_zaznamov": imported_rows,
        "pocet_odstranenych_duplicit": removed_duplicates,
        "pocet_vyradenych_bez_emailu_a_telefonu": dropped_without_channel,
        "pocet_finalnych_kontaktov": len(deduplicated_contacts),
    }
    write_report(report_path, report)

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
