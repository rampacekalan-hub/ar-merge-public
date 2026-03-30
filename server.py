#!/usr/bin/env python3
import csv
import base64
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import sqlite3
import smtplib
import socket
from datetime import date, datetime, timedelta, timezone
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib import error as urllib_error
from urllib import request as urllib_request
from urllib.parse import parse_qs, urlparse

import openpyxl
import stripe
import xlrd
import fitz
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from contact_importer import OUTPUT_HEADERS, process_uploads
from file_compressor import compress_upload


HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", "8080"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "app.db"))
SESSION_COOKIE_NAME = "cdc_session"
SESSION_TTL_SECONDS = 60 * 60
PASSWORD_ITERATIONS = 200_000
STRIPE_SECRET_KEY = os.environ.get("STRIPE_SECRET_KEY", "").strip()
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET", "").strip()
STRIPE_PRICE_ID = os.environ.get("STRIPE_PRICE_ID", "").strip()
DEFAULT_STRIPE_PRICE_ID = os.environ.get("DEFAULT_STRIPE_PRICE_ID", "price_1TG0m1JJsHlQYfgdgMnHLJ8r").strip()
APP_BASE_URL = os.environ.get("APP_BASE_URL", "https://unifyo.online").strip()
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-5-mini").strip() or "gpt-5-mini"
SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "").strip()
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "").strip()
SMTP_FROM_EMAIL = os.environ.get("SMTP_FROM_EMAIL", "").strip()
SUPPORT_EMAIL = "info@unifyo.online"
ADMIN_EMAILS = {item.strip().lower() for item in os.environ.get("ADMIN_EMAILS", "").split(",") if item.strip()}
MAX_REQUEST_BODY_MB = float(os.environ.get("MAX_REQUEST_BODY_MB", "110"))
MAX_CONTACT_FILE_MB = float(os.environ.get("MAX_CONTACT_FILE_MB", "8"))
MAX_COMPRESS_FILE_MB = float(os.environ.get("MAX_COMPRESS_FILE_MB", "100"))
OPENAI_WEB_SEARCH_ENABLED = os.environ.get("OPENAI_ENABLE_WEB_SEARCH", "1").strip().lower() not in {"0", "false", "no"}
OPENAI_WEB_SEARCH_TOOL = os.environ.get("OPENAI_WEB_SEARCH_TOOL", "web_search_preview").strip() or "web_search_preview"
OPENAI_WEB_SEARCH_RUNTIME_DISABLED = False
MAX_AI_IMAGE_MB = float(os.environ.get("MAX_AI_IMAGE_MB", "10"))
PROMPT_VERSION = "unifyo-sk-fin-v4"
LEGAL_VERSION = "2026-03-27"
CHECKOUT_CONSENT_VERSION = "2026-03-29"
MEMBERSHIP_SYNC_COOLDOWN_SECONDS = max(5, int(os.environ.get("MEMBERSHIP_SYNC_COOLDOWN_SECONDS", "45")))
ADMIN_SYNC_COOLDOWN_SECONDS = max(30, int(os.environ.get("ADMIN_SYNC_COOLDOWN_SECONDS", "300")))
REGISTRATION_CONFIRMATION_TEXT = (
    "Výberom možnosti Vytvoriť účet potvrdzujete, že máte minimálne 18 rokov, "
    "súhlasíte s Obchodnými podmienkami a zároveň potvrdzujete, že ste si prečítali "
    "naše Zásady ochrany súkromia - GDPR."
)
CHECKOUT_CONSENT_TEXT = (
    "Súhlasím so začatím poskytovania služby pred uplynutím lehoty na odstúpenie od zmluvy "
    "a beriem na vedomie, že tým strácam právo na odstúpenie od zmluvy."
)
TRUSTED_SOURCE_DOMAINS = {
    "nbs.sk",
    "slov-lex.sk",
    "fininfo.sk",
    "ecb.europa.eu",
    "europa.eu",
    "minv.sk",
    "finance.gov.sk",
    "alanrampacek.sk",
    "prosight.sk",
}

MAX_REQUEST_BODY_BYTES = max(1, int(MAX_REQUEST_BODY_MB * 1024 * 1024))
MAX_CONTACT_FILE_BYTES = max(1, int(MAX_CONTACT_FILE_MB * 1024 * 1024))
MAX_COMPRESS_FILE_BYTES = max(1, int(MAX_COMPRESS_FILE_MB * 1024 * 1024))
MAX_AI_IMAGE_BYTES = max(1, int(MAX_AI_IMAGE_MB * 1024 * 1024))

stripe.api_key = STRIPE_SECRET_KEY
LAST_MEMBERSHIP_SYNC_BY_USER = {}
LAST_ADMIN_SYNC_AT = None


def resolve_stripe_price_id():
    raw = (STRIPE_PRICE_ID or "").strip()
    if re.fullmatch(r"price_[A-Za-z0-9]+", raw):
        return raw
    if raw:
        match = re.search(r"(price_[A-Za-z0-9]+)", raw)
        if match:
            return match.group(1)
    if re.fullmatch(r"price_[A-Za-z0-9]+", DEFAULT_STRIPE_PRICE_ID or ""):
        return DEFAULT_STRIPE_PRICE_ID
    return ""


def utc_now():
    return datetime.now(timezone.utc)


def format_timestamp(value):
    if isinstance(value, str):
        return value
    if not value:
        return ""
    return value.astimezone(timezone.utc).isoformat()


def parse_timestamp(value):
    if not value:
        return None
    return datetime.fromisoformat(value)


def get_db():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def init_db():
    db_dir = os.path.dirname(DB_PATH)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    connection = get_db()
    try:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                password_salt TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                stripe_customer_id TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                expires_at TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS memberships (
                user_id INTEGER PRIMARY KEY,
                status TEXT NOT NULL,
                stripe_customer_id TEXT,
                stripe_subscription_id TEXT,
                stripe_status TEXT NOT NULL DEFAULT '',
                cancel_at_period_end INTEGER NOT NULL DEFAULT 0,
                cancelled_at TEXT NOT NULL DEFAULT '',
                current_period_end TEXT,
                next_renewal_at TEXT NOT NULL DEFAULT '',
                internal_subscription_number TEXT NOT NULL DEFAULT '',
                last_order_number TEXT NOT NULL DEFAULT '',
                last_checkout_session_id TEXT NOT NULL DEFAULT '',
                last_payment_intent_id TEXT NOT NULL DEFAULT '',
                last_invoice_id TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS checkout_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                stripe_session_id TEXT NOT NULL UNIQUE,
                internal_order_number TEXT NOT NULL DEFAULT '',
                checkout_consent_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS registration_consents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                email TEXT NOT NULL,
                accepted_terms INTEGER NOT NULL DEFAULT 0,
                accepted_privacy INTEGER NOT NULL DEFAULT 0,
                marketing_consent INTEGER NOT NULL DEFAULT 0,
                consent_text TEXT NOT NULL DEFAULT '',
                legal_version TEXT NOT NULL DEFAULT '',
                ip_address TEXT NOT NULL DEFAULT '',
                user_agent TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS checkout_consents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                email TEXT NOT NULL,
                consent_text TEXT NOT NULL DEFAULT '',
                consent_version TEXT NOT NULL DEFAULT '',
                price_label TEXT NOT NULL DEFAULT '',
                subscription_label TEXT NOT NULL DEFAULT '',
                renewal_label TEXT NOT NULL DEFAULT '',
                no_refund_label TEXT NOT NULL DEFAULT '',
                ip_address TEXT NOT NULL DEFAULT '',
                user_agent TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS subscription_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                event_type TEXT NOT NULL,
                stripe_customer_id TEXT NOT NULL DEFAULT '',
                stripe_subscription_id TEXT NOT NULL DEFAULT '',
                stripe_checkout_session_id TEXT NOT NULL DEFAULT '',
                stripe_payment_intent_id TEXT NOT NULL DEFAULT '',
                stripe_invoice_id TEXT NOT NULL DEFAULT '',
                internal_order_number TEXT NOT NULL DEFAULT '',
                internal_subscription_number TEXT NOT NULL DEFAULT '',
                subscription_status TEXT NOT NULL DEFAULT '',
                cancel_at_period_end INTEGER NOT NULL DEFAULT 0,
                next_renewal_at TEXT NOT NULL DEFAULT '',
                access_until TEXT NOT NULL DEFAULT '',
                event_meta TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                expires_at TEXT NOT NULL,
                created_at TEXT NOT NULL,
                used_at TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS activity_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                event_type TEXT NOT NULL,
                event_label TEXT NOT NULL,
                event_meta TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS assistant_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                client_name TEXT NOT NULL DEFAULT '',
                due_date TEXT NOT NULL DEFAULT '',
                priority TEXT NOT NULL DEFAULT 'medium',
                channel TEXT NOT NULL DEFAULT 'followup',
                details TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'open',
                completed_at TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS assistant_profiles (
                user_id INTEGER PRIMARY KEY,
                focus TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS assistant_threads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL DEFAULT 'Nový chat',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS assistant_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                thread_id INTEGER,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                meta_json TEXT NOT NULL DEFAULT '{}',
                review_status TEXT NOT NULL DEFAULT 'unreviewed',
                reviewed_by INTEGER,
                reviewed_at TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY (thread_id) REFERENCES assistant_threads(id) ON DELETE CASCADE,
                FOREIGN KEY (reviewed_by) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS generated_assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token TEXT NOT NULL UNIQUE,
                file_name TEXT NOT NULL,
                content_type TEXT NOT NULL,
                payload_base64 TEXT NOT NULL,
                kind TEXT NOT NULL DEFAULT 'file',
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS assistant_upload_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'pending',
                attachment_name TEXT NOT NULL DEFAULT '',
                attachment_type TEXT NOT NULL DEFAULT '',
                attachment_base64 TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                uploaded_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            """
        )
        columns = {row["name"] for row in connection.execute("PRAGMA table_info(users)").fetchall()}
        if "name" not in columns:
            connection.execute("ALTER TABLE users ADD COLUMN name TEXT NOT NULL DEFAULT ''")
        if "role" not in columns:
            connection.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'user'")
        membership_columns = {row["name"] for row in connection.execute("PRAGMA table_info(memberships)").fetchall()}
        if "stripe_status" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN stripe_status TEXT NOT NULL DEFAULT ''")
        if "cancel_at_period_end" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN cancel_at_period_end INTEGER NOT NULL DEFAULT 0")
        if "cancelled_at" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN cancelled_at TEXT NOT NULL DEFAULT ''")
        if "next_renewal_at" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN next_renewal_at TEXT NOT NULL DEFAULT ''")
        if "internal_subscription_number" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN internal_subscription_number TEXT NOT NULL DEFAULT ''")
        if "last_order_number" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN last_order_number TEXT NOT NULL DEFAULT ''")
        if "last_checkout_session_id" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN last_checkout_session_id TEXT NOT NULL DEFAULT ''")
        if "last_payment_intent_id" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN last_payment_intent_id TEXT NOT NULL DEFAULT ''")
        if "last_invoice_id" not in membership_columns:
            connection.execute("ALTER TABLE memberships ADD COLUMN last_invoice_id TEXT NOT NULL DEFAULT ''")
        checkout_columns = {row["name"] for row in connection.execute("PRAGMA table_info(checkout_sessions)").fetchall()}
        if "internal_order_number" not in checkout_columns:
            connection.execute("ALTER TABLE checkout_sessions ADD COLUMN internal_order_number TEXT NOT NULL DEFAULT ''")
        if "checkout_consent_id" not in checkout_columns:
            connection.execute("ALTER TABLE checkout_sessions ADD COLUMN checkout_consent_id INTEGER")
        assistant_message_columns = {
            row["name"] for row in connection.execute("PRAGMA table_info(assistant_messages)").fetchall()
        }
        if "thread_id" not in assistant_message_columns:
            connection.execute("ALTER TABLE assistant_messages ADD COLUMN thread_id INTEGER")
        if "meta_json" not in assistant_message_columns:
            connection.execute("ALTER TABLE assistant_messages ADD COLUMN meta_json TEXT NOT NULL DEFAULT '{}'")
        if "review_status" not in assistant_message_columns:
            connection.execute("ALTER TABLE assistant_messages ADD COLUMN review_status TEXT NOT NULL DEFAULT 'unreviewed'")
        if "reviewed_by" not in assistant_message_columns:
            connection.execute("ALTER TABLE assistant_messages ADD COLUMN reviewed_by INTEGER")
        if "reviewed_at" not in assistant_message_columns:
            connection.execute("ALTER TABLE assistant_messages ADD COLUMN reviewed_at TEXT NOT NULL DEFAULT ''")

        orphan_user_ids = [
            row["user_id"]
            for row in connection.execute(
                """
                SELECT DISTINCT user_id
                FROM assistant_messages
                WHERE thread_id IS NULL OR thread_id = 0
                """
            ).fetchall()
        ]
        for user_id in orphan_user_ids:
            now_value = format_timestamp(utc_now())
            cursor = connection.execute(
                """
                INSERT INTO assistant_threads (user_id, title, created_at, updated_at)
                VALUES (?, 'Starší chat', ?, ?)
                """,
                (user_id, now_value, now_value),
            )
            connection.execute(
                """
                UPDATE assistant_messages
                SET thread_id = ?
                WHERE user_id = ? AND (thread_id IS NULL OR thread_id = 0)
                """,
                (cursor.lastrowid, user_id),
            )
        connection.commit()
    finally:
        connection.close()


def normalize_email(email):
    return str(email or "").strip().lower()


def parse_positive_int(value, default=0):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed


def format_mb(size_bytes):
    return f"{size_bytes / (1024 * 1024):.2f} MB"


def safe_json_loads(value, default=None):
    if default is None:
        default = {}
    try:
        return json.loads(value or "")
    except Exception:
        return default


def today_iso():
    return datetime.now().astimezone().date().isoformat()


def is_admin_email(email):
    return normalize_email(email) in ADMIN_EMAILS


def can_manage_admin_tools(user_row):
    return bool(user_row) and user_row["role"] == "admin" and is_admin_email(user_row["email"])


def hash_password(password, salt):
    return hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        bytes.fromhex(salt),
        PASSWORD_ITERATIONS,
    ).hex()


def create_password_hash(password):
    salt = secrets.token_hex(16)
    password_hash = hash_password(password, salt)
    return salt, password_hash


def verify_password(password, salt, password_hash):
    return hmac.compare_digest(hash_password(password, salt), password_hash)


def hash_session_token(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def build_reset_link(token):
    base = APP_BASE_URL.rstrip("/") if APP_BASE_URL else ""
    return f"{base}/app.html?reset_token={token}"


def send_reset_email(to_email, name, token):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_FROM_EMAIL):
        raise RuntimeError("SMTP nie je nakonfigurované pre obnovu hesla.")

    reset_link = build_reset_link(token)
    message = EmailMessage()
    message["Subject"] = "Obnova hesla - Unifyo"
    message["From"] = SMTP_FROM_EMAIL
    message["To"] = to_email
    greeting = name or to_email
    message.set_content(
        f"""Dobrý deň {greeting},

prišla nám žiadosť o obnovu hesla pre váš účet v Unifyo.

Pre nastavenie nového hesla otvorte tento odkaz:
{reset_link}

Platnosť odkazu je 30 minút.

Ak ste o obnovu hesla nežiadali, tento e-mail môžete ignorovať.
"""
    )

    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
            smtp.login(SMTP_USER, SMTP_PASSWORD)
            smtp.send_message(message)
        return

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(message)


def send_contact_email(sender_name, sender_email, subject, message_text):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_FROM_EMAIL):
        raise RuntimeError("SMTP nie je nakonfigurované pre kontaktný formulár.")

    message = EmailMessage()
    message["Subject"] = f"Unifyo kontakt: {subject}"
    message["From"] = SMTP_FROM_EMAIL
    message["To"] = SMTP_FROM_EMAIL
    message["Reply-To"] = sender_email
    message.set_content(
        f"""Nová správa z kontaktného formulára Unifyo

Meno / kontakt:
{sender_name}

E-mail:
{sender_email}

Predmet:
{subject}

Správa:
{message_text}
"""
    )

    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
            smtp.login(SMTP_USER, SMTP_PASSWORD)
            smtp.send_message(message)
        return

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(message)


def send_account_deleted_email(to_email, name):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_FROM_EMAIL):
        return

    message = EmailMessage()
    message["Subject"] = "Unifyo - účet bol zrušený"
    message["From"] = SMTP_FROM_EMAIL
    message["To"] = to_email
    greeting = name or to_email
    message.set_content(
        f"""Dobrý deň {greeting},

váš účet v Unifyo bol administratívne zrušený.

Ak na účte prebehla platba alebo obnova členstva, prípadné vrátenie platby bude spracované do 7 dní.

Ak ide o nedorozumenie, odpovedzte priamo na tento e-mail.
"""
    )

    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
            smtp.login(SMTP_USER, SMTP_PASSWORD)
            smtp.send_message(message)
        return

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(message)


def send_subscription_email(to_email, name, subject, body_text, bcc_support=True):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_FROM_EMAIL):
        return

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = SMTP_FROM_EMAIL
    message["To"] = to_email
    if bcc_support and SUPPORT_EMAIL:
        message["Bcc"] = SUPPORT_EMAIL
    greeting = name or to_email
    message.set_content(f"Dobrý deň {greeting},\n\n{body_text}\n")

    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
            smtp.login(SMTP_USER, SMTP_PASSWORD)
            smtp.send_message(message)
        return

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(message)


def send_registration_welcome_email(user_row):
    body = f"""účet UNIFYO bol úspešne vytvorený.

Meno: {user_row.get('name') or '—'}
E-mail: {user_row.get('email') or '—'}
Dátum registrácie: {format_timestamp(utc_now())}

V účte môžeš okamžite:
- aktivovať mesačné členstvo
- používať čistenie kontaktov, kompresiu súborov a AI asistenta

Podpora: {SUPPORT_EMAIL}
"""
    send_subscription_email(
        user_row["email"],
        user_row["name"],
        "UNIFYO - potvrdenie registrácie",
        body,
        bcc_support=True,
    )


def send_subscription_activated_email(user_row, details):
    body = f"""ďakujeme, predplatné UNIFYO bolo úspešne aktivované.

Mesačné predplatné: {details.get('price_label') or '1,99 € / mesiac'}
Automatické obnovenie: áno
Dátum nákupu: {details.get('purchased_at') or ''}
Ďalšie obnovenie: {details.get('next_renewal_at') or ''}
Interné číslo objednávky: {details.get('internal_order_number') or '—'}
Interné číslo predplatného: {details.get('internal_subscription_number') or '—'}
Stripe subscription ID: {details.get('stripe_subscription_id') or '—'}
Stripe payment ID: {details.get('stripe_payment_intent_id') or '—'}

Predplatné môžeš kedykoľvek zrušiť priamo vo svojom účte alebo e-mailom. Po zrušení zostane prístup aktívny do konca už zaplateného obdobia a ďalšia platba sa už nestrhne.

Podpora: {SUPPORT_EMAIL}
"""
    send_subscription_email(
        user_row["email"],
        user_row["name"],
        "UNIFYO - potvrdenie aktivácie predplatného",
        body,
        bcc_support=True,
    )


def send_subscription_cancelled_email(user_row, details):
    body = f"""potvrdzujeme zrušenie predplatného UNIFYO.

Prístup zostáva aktívny do: {details.get('access_until') or ''}
Ďalšia platba už nebude účtovaná.
Interné číslo objednávky: {details.get('internal_order_number') or '—'}
Interné číslo predplatného: {details.get('internal_subscription_number') or '—'}
Stripe subscription ID: {details.get('stripe_subscription_id') or '—'}

Ak potrebuješ pomoc, kontaktuj nás na {SUPPORT_EMAIL}.
"""
    send_subscription_email(
        user_row["email"],
        user_row["name"],
        "UNIFYO - potvrdenie zrušenia predplatného",
        body,
        bcc_support=True,
    )


def generate_internal_order_number():
    now_value = datetime.now().astimezone().strftime("%Y%m%d")
    return f"UNY-OBJ-{now_value}-{secrets.token_hex(3).upper()}"


def generate_internal_subscription_number():
    now_value = datetime.now().astimezone().strftime("%Y%m%d")
    return f"UNY-SUB-{now_value}-{secrets.token_hex(3).upper()}"


def record_registration_consent(connection, user_id, email, ip_address, user_agent, marketing_consent):
    now_value = format_timestamp(utc_now())
    cursor = connection.execute(
        """
        INSERT INTO registration_consents (
            user_id, email, ip_address, accepted_terms, accepted_privacy, marketing_consent,
            consent_text, legal_version, user_agent, created_at
        ) VALUES (?, ?, ?, 1, 1, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            email,
            ip_address,
            1 if marketing_consent else 0,
            REGISTRATION_CONFIRMATION_TEXT,
            LEGAL_VERSION,
            user_agent,
            now_value,
        ),
    )
    return cursor.lastrowid


def record_checkout_consent(connection, user_row, ip_address, user_agent):
    now_value = format_timestamp(utc_now())
    cursor = connection.execute(
        """
        INSERT INTO checkout_consents (
            user_id, email, consent_text, consent_version, price_label, subscription_label,
            renewal_label, no_refund_label, ip_address, user_agent, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_row["id"],
            user_row["email"],
            CHECKOUT_CONSENT_TEXT,
            CHECKOUT_CONSENT_VERSION,
            "1,99 € / mesiac",
            "Mesačné predplatné s automatickým obnovením",
            "Predplatné sa automaticky obnovuje, kým ho nezrušíte.",
            "Platba sa po aktivácii služby nevracia.",
            ip_address,
            user_agent,
            now_value,
        ),
    )
    return cursor.lastrowid


def record_subscription_event(connection, user_id, event_type, status, details=None):
    now_value = format_timestamp(utc_now())
    connection.execute(
        """
        INSERT INTO subscription_events (user_id, event_type, status, details_json, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            user_id,
            event_type,
            status,
            json.dumps(details or {}, ensure_ascii=False),
            now_value,
        ),
    )


def get_latest_registration_consent(connection, user_id):
    row = connection.execute(
        """
        SELECT *
        FROM registration_consents
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id,),
    ).fetchone()
    return dict(row) if row else None


def get_latest_checkout_consent(connection, user_id):
    row = connection.execute(
        """
        SELECT *
        FROM checkout_consents
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id,),
    ).fetchone()
    return dict(row) if row else None


def sanitize_download_filename(value, fallback="export"):
    raw = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    raw = raw.strip("._-")
    return raw[:80] or fallback


def parse_markdown_table(text):
    lines = [line.rstrip() for line in str(text or "").splitlines() if line.strip()]
    for index in range(len(lines) - 1):
        header_line = lines[index]
        divider_line = lines[index + 1]
        if "|" not in header_line or "|" not in divider_line:
            continue
        divider_clean = divider_line.replace("|", "").replace(":", "").replace("-", "").replace(" ", "")
        if divider_clean:
            continue
        headers = [cell.strip() for cell in header_line.strip().strip("|").split("|")]
        rows = []
        for row_line in lines[index + 2:]:
            if "|" not in row_line:
                break
            values = [cell.strip() for cell in row_line.strip().strip("|").split("|")]
            if len(values) != len(headers):
                break
            rows.append(values)
        if headers and rows:
            return {"headers": headers, "rows": rows}
    return None


def build_pdf_bytes(title, body_text, table_data=None):
    document = fitz.open()
    page = document.new_page()
    margin_x = 46
    y = 54
    page.insert_text((margin_x, y), title or "Unifyo export", fontsize=18, fontname="helv", fill=(0.07, 0.13, 0.24))
    y += 28

    if table_data and table_data.get("headers") and table_data.get("rows"):
        headers = table_data["headers"][:6]
        rows = table_data["rows"][:24]
        widths = [78, 86, 82, 96, 96, 96][: len(headers)]
        x_positions = [margin_x]
        for width in widths[:-1]:
            x_positions.append(x_positions[-1] + width)

        for idx, header in enumerate(headers):
            page.insert_text((x_positions[idx], y), str(header)[:28], fontsize=10, fontname="helv", fill=(0.1, 0.22, 0.4))
        y += 18
        for row in rows:
            for idx, value in enumerate(row[: len(headers)]):
                page.insert_text((x_positions[idx], y), str(value)[:30], fontsize=9, fontname="helv", fill=(0.15, 0.18, 0.24))
            y += 16
            if y > 760:
                page = document.new_page()
                y = 54
    else:
        text = str(body_text or "").strip()
        for paragraph in [item.strip() for item in text.split("\n") if item.strip()]:
            rect = fitz.Rect(margin_x, y, 550, 790)
            written = page.insert_textbox(rect, paragraph, fontsize=11, fontname="helv", fill=(0.15, 0.18, 0.24), lineheight=1.25)
            y += max(26, int((rect.height - written) / 8) if written < 0 else 24)
            if y > 760:
                page = document.new_page()
                y = 54

    pdf_bytes = document.tobytes()
    document.close()
    return pdf_bytes


def strip_markdown_table_lines(text):
    lines = str(text or "").splitlines()
    kept = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            continue
        kept.append(line)
    return "\n".join(kept).strip()


def should_generate_pdf_asset(user_message, assistant_reply):
    combined = f"{user_message or ''} {assistant_reply or ''}".lower()
    triggers = (
        " pdf",
        "v pdf",
        "do pdf",
        "stiahnuť pdf",
        "stiahnut pdf",
        "download pdf",
        "pdf-ready",
    )
    return any(trigger in combined for trigger in triggers)


def store_generated_asset(connection, user_id, file_name, content_type, payload_bytes, kind="file", ttl_minutes=30):
    now = utc_now()
    token = secrets.token_urlsafe(18)
    connection.execute(
        """
        INSERT INTO generated_assets (user_id, token, file_name, content_type, payload_base64, kind, created_at, expires_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            token,
            file_name,
            content_type,
            base64.b64encode(payload_bytes).decode("ascii"),
            kind,
            format_timestamp(now),
            format_timestamp(now + timedelta(minutes=ttl_minutes)),
        ),
    )
    connection.commit()
    return token


def get_generated_asset(connection, user_id, token):
    row = connection.execute(
        "SELECT * FROM generated_assets WHERE token = ? AND user_id = ?",
        (str(token or "").strip(), user_id),
    ).fetchone()
    if not row:
        return None
    expires_at = parse_timestamp(row["expires_at"])
    if not expires_at or expires_at <= utc_now():
        connection.execute("DELETE FROM generated_assets WHERE token = ?", (token,))
        connection.commit()
        return None
    return row


def require_json(handler):
    try:
        content_length = int(handler.headers.get("Content-Length", "0"))
        payload = json.loads(handler.rfile.read(content_length).decode("utf-8"))
    except Exception:
        raise ValueError("Neplatné JSON dáta.")
    return payload


def get_cookie_token(headers):
    cookie_header = headers.get("Cookie", "")
    if not cookie_header:
        return ""
    cookie = SimpleCookie()
    cookie.load(cookie_header)
    morsel = cookie.get(SESSION_COOKIE_NAME)
    return morsel.value if morsel else ""


def create_session(connection, user_id):
    token = secrets.token_urlsafe(32)
    now = utc_now()
    expires_at = now + timedelta(seconds=SESSION_TTL_SECONDS)
    connection.execute(
        """
        INSERT INTO sessions (user_id, token_hash, expires_at, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (user_id, hash_session_token(token), format_timestamp(expires_at), format_timestamp(now)),
    )
    connection.commit()
    return token, expires_at


def log_activity(connection, event_type, event_label, user_id=None, **meta):
    connection.execute(
        """
        INSERT INTO activity_logs (user_id, event_type, event_label, event_meta, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            user_id,
            event_type,
            event_label,
            json.dumps(meta, ensure_ascii=False),
            format_timestamp(utc_now()),
        ),
    )
    connection.commit()


def get_request_ip(handler):
    forwarded = str(handler.headers.get("X-Forwarded-For", "") or "").strip()
    if forwarded:
        return forwarded.split(",")[0].strip()
    client = getattr(handler, "client_address", None)
    if client and client[0]:
        return str(client[0])
    return ""


def get_request_agent(handler):
    return str(handler.headers.get("User-Agent", "") or "").strip()


def clear_session(connection, token):
    if not token:
        return
    connection.execute("DELETE FROM sessions WHERE token_hash = ?", (hash_session_token(token),))
    connection.commit()


def get_session_user(connection, headers):
    token = get_cookie_token(headers)
    if not token:
        return None

    row = connection.execute(
        """
        SELECT users.*
        FROM sessions
        JOIN users ON users.id = sessions.user_id
        WHERE sessions.token_hash = ?
        """,
        (hash_session_token(token),),
    ).fetchone()
    if not row:
        return None

    session_row = connection.execute(
        "SELECT expires_at FROM sessions WHERE token_hash = ?",
        (hash_session_token(token),),
    ).fetchone()
    expires_at = parse_timestamp(session_row["expires_at"]) if session_row else None
    if not expires_at or expires_at <= utc_now():
        clear_session(connection, token)
        return None
    return row


def get_session_expiry(connection, headers):
    token = get_cookie_token(headers)
    if not token:
        return ""
    row = connection.execute(
        "SELECT expires_at FROM sessions WHERE token_hash = ?",
        (hash_session_token(token),),
    ).fetchone()
    return row["expires_at"] if row else ""


def get_membership(connection, user_id):
    row = connection.execute(
        "SELECT * FROM memberships WHERE user_id = ?",
        (user_id,),
    ).fetchone()
    if not row:
        return None
    membership = dict(row)
    membership["current_period_end"] = parse_timestamp(membership.get("current_period_end"))
    membership["created_at"] = parse_timestamp(membership.get("created_at"))
    membership["updated_at"] = parse_timestamp(membership.get("updated_at"))
    membership["cancelled_at"] = parse_timestamp(membership.get("cancelled_at"))
    membership["next_renewal_at"] = parse_timestamp(membership.get("next_renewal_at"))
    membership["cancel_at_period_end"] = bool(membership.get("cancel_at_period_end"))
    return membership


def ensure_admin_role(connection, user_row):
    if not user_row:
        return user_row
    if user_row["role"] == "admin":
        return user_row
    if not is_admin_email(user_row["email"]):
        return user_row
    connection.execute(
        "UPDATE users SET role = 'admin', updated_at = ? WHERE id = ?",
        (format_timestamp(utc_now()), user_row["id"]),
    )
    connection.commit()
    return connection.execute("SELECT * FROM users WHERE id = ?", (user_row["id"],)).fetchone()


def get_recent_activity(connection, user_id=None, limit=20):
    params = []
    query = """
        SELECT activity_logs.*, users.email AS user_email, users.name AS user_name
        FROM activity_logs
        LEFT JOIN users ON users.id = activity_logs.user_id
    """
    if user_id is not None:
        query += " WHERE activity_logs.user_id = ?"
        params.append(user_id)
    query += " ORDER BY activity_logs.created_at DESC LIMIT ?"
    params.append(limit)
    rows = connection.execute(query, tuple(params)).fetchall()
    activity = []
    for row in rows:
        try:
            meta = json.loads(row["event_meta"] or "{}")
        except Exception:
            meta = {}
        activity.append(
            {
                "id": row["id"],
                "user_id": row["user_id"],
                "user_email": row["user_email"] or "",
                "user_name": row["user_name"] or "",
                "event_type": row["event_type"],
                "event_label": row["event_label"],
                "created_at": row["created_at"],
                "meta": meta,
            }
        )
    return activity


def get_recent_subscription_events(connection, user_id=None, limit=12):
    params = []
    query = """
        SELECT *
        FROM subscription_events
    """
    if user_id is not None:
        query += " WHERE user_id = ?"
        params.append(user_id)
    query += " ORDER BY created_at DESC, id DESC LIMIT ?"
    params.append(limit)
    rows = connection.execute(query, tuple(params)).fetchall()
    return [
        {
            "id": row["id"],
            "user_id": row["user_id"],
            "event_type": row["event_type"],
            "status": row["status"],
            "details": safe_json_loads(row["details_json"], {}),
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def get_subscription_snapshot(connection, user_id):
    membership = get_membership(connection, user_id) or {}
    registration_consent = get_latest_registration_consent(connection, user_id) or {}
    checkout_consent = get_latest_checkout_consent(connection, user_id) or {}
    return {
        "membership": {
            "status": membership.get("status") or "inactive",
            "stripe_status": membership.get("stripe_status") or "",
            "valid_until": format_timestamp(membership.get("current_period_end")) if membership.get("current_period_end") else "",
            "started_at": format_timestamp(membership.get("created_at")) if membership.get("created_at") else "",
            "next_renewal_at": format_timestamp(membership.get("next_renewal_at")) if membership.get("next_renewal_at") else "",
            "cancelled_at": format_timestamp(membership.get("cancelled_at")) if membership.get("cancelled_at") else "",
            "cancel_at_period_end": bool(membership.get("cancel_at_period_end")),
            "internal_subscription_number": membership.get("internal_subscription_number") or "",
            "last_order_number": membership.get("last_order_number") or "",
            "last_checkout_session_id": membership.get("last_checkout_session_id") or "",
            "last_payment_intent_id": membership.get("last_payment_intent_id") or "",
            "last_invoice_id": membership.get("last_invoice_id") or "",
            "stripe_subscription_id": membership.get("stripe_subscription_id") or "",
        },
        "registration_consent": {
            "created_at": registration_consent.get("created_at") or "",
            "ip_address": registration_consent.get("ip_address") or "",
            "marketing_consent": bool(registration_consent.get("marketing_consent")),
            "legal_version": registration_consent.get("legal_version") or "",
        },
        "checkout_consent": {
            "created_at": checkout_consent.get("created_at") or "",
            "ip_address": checkout_consent.get("ip_address") or "",
            "consent_version": checkout_consent.get("consent_version") or "",
            "consent_text": checkout_consent.get("consent_text") or "",
        },
        "events": get_recent_subscription_events(connection, user_id=user_id, limit=12),
    }


def row_to_task(row):
    return {
        "id": row["id"],
        "title": row["title"],
        "client_name": row["client_name"] or "",
        "due_date": row["due_date"] or "",
        "priority": row["priority"] or "medium",
        "channel": row["channel"] or "followup",
        "details": row["details"] or "",
        "status": row["status"] or "open",
        "completed_at": row["completed_at"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def get_assistant_tasks(connection, user_id, include_done=True):
    query = """
        SELECT *
        FROM assistant_tasks
        WHERE user_id = ?
    """
    params = [user_id]
    if not include_done:
        query += " AND status != 'done'"
    query += """
        ORDER BY
            CASE priority
                WHEN 'high' THEN 0
                WHEN 'medium' THEN 1
                ELSE 2
            END,
            CASE
                WHEN due_date = '' THEN 1
                ELSE 0
            END,
            due_date ASC,
            created_at DESC
    """
    return [row_to_task(row) for row in connection.execute(query, tuple(params)).fetchall()]


def build_assistant_brief(tasks):
    today = today_iso()
    open_tasks = [task for task in tasks if task["status"] != "done"]
    done_tasks = [task for task in tasks if task["status"] == "done"]
    overdue = [task for task in open_tasks if task["due_date"] and task["due_date"] < today]
    due_today = [task for task in open_tasks if task["due_date"] == today]
    high_priority = [task for task in open_tasks if task["priority"] == "high"]

    counts = {
        "open_tasks": len(open_tasks),
        "due_today": len(due_today),
        "overdue": len(overdue),
        "completed": len(done_tasks),
        "high_priority": len(high_priority),
    }

    focus = []
    if overdue:
        focus.append(f"Najprv vybav {len(overdue)} omeškaných úloh.")
    if due_today:
        focus.append(f"Dnes máš {len(due_today)} úloh na follow-up alebo kontakt.")
    if high_priority:
        focus.append(f"Prioritu má {len(high_priority)} dôležitých klientskych krokov.")
    if not focus:
        focus.append("Dnes máš priestor pripraviť nové follow-upy a upratať pipeline.")

    suggestions = []
    top_tasks = open_tasks[:3]
    for task in top_tasks:
        client = task["client_name"] or "klienta"
        if task["channel"] == "email":
            suggestions.append({
                "title": f"E-mail follow-up: {client}",
                "body": f"Dobrý deň {client}, ozývam sa k nášmu poslednému kontaktu. Navrhujem krátke doplnenie podkladov a dohodnutie ďalšieho kroku ešte dnes.",
            })
        elif task["channel"] == "call":
            suggestions.append({
                "title": f"Call script: {client}",
                "body": f"Dnes kontaktuj {client} telefonicky, over stav rozpracovania a navrhni konkrétny termín ďalšieho kroku.",
            })
        else:
            suggestions.append({
                "title": f"Ďalší krok: {client}",
                "body": f"Skontroluj úlohu „{task['title']}“ a posuň klienta {client} do ďalšej fázy bez zbytočného odkladu.",
            })

    if not suggestions:
        suggestions.append({
            "title": "Denný štart",
            "body": "Začni kontrolou nových leadov, obnov termíny follow-upov a priprav si tri najdôležitejšie klientské kroky na dnes.",
        })

    return {
        "today": today,
        "counts": counts,
        "focus": focus,
        "suggestions": suggestions,
    }


def get_assistant_profile(connection, user_id):
    row = connection.execute(
        "SELECT * FROM assistant_profiles WHERE user_id = ?",
        (user_id,),
    ).fetchone()
    if not row:
        return {"focus": "", "notes": ""}
    return {
        "focus": row["focus"] or "",
        "notes": row["notes"] or "",
        "updated_at": row["updated_at"] or "",
    }


def build_thread_title(message, language="sk"):
    text = " ".join(str(message or "").strip().split())
    if not text:
        return "New chat" if language == "en" else "Nový chat"
    text = re.sub(r"\[priložený obrázok:[^\]]+\]", "", text, flags=re.IGNORECASE).strip()
    if not text:
        return "New chat" if language == "en" else "Nový chat"
    lowered = text.lower()
    if any(token in lowered for token in ("úrok", "urok", "sadzb", "fixáci", "fixac", "hypot")) and any(
        token in lowered for token in ("banka", "bank", "slovensk", "slovensku", "svk")
    ):
        return "Bank rates in Slovakia" if language == "en" else "Úroky v bankách na Slovensku"
    if any(token in lowered for token in ("inflác", "inflac", "deflác", "deflac", "ceny", "zdraž")):
        return "Inflation and prices in Slovakia" if language == "en" else "Inflácia a ceny na Slovensku"
    if any(token in lowered for token in ("poisten", "krytie", "rizik")):
        return "Insurance coverage for clients" if language == "en" else "Poistenie a krytie klienta"
    if any(token in lowered for token in ("invest", "portfól", "portfolio", "fond")):
        return "Investing and client portfolio" if language == "en" else "Investovanie a portfólio klienta"
    if any(token in lowered for token in ("refinanc", "refi")):
        return "Mortgage refinancing in Slovakia" if language == "en" else "Refinancovanie hypotéky na Slovensku"
    first_sentence = re.split(r"[.!?\n]", text, maxsplit=1)[0].strip()
    first_sentence = re.sub(
        r"^(prosím|prosim|ahoj|dobrý deň|dobry den|čau|cau|hello|hi)\s*[,:-]*\s*",
        "",
        first_sentence,
        flags=re.IGNORECASE,
    ).strip()
    first_sentence = re.sub(
        r"^(priprav mi|napíš mi|napis mi|pomôž mi|pomoz mi|vyhodnoť|vyhodnot|vysvetli|navrhni|zhrň|zhrn|zjednoduš|zjednodus|preformuluj|napíš|napis)\s+",
        "",
        first_sentence,
        flags=re.IGNORECASE,
    ).strip()
    first_sentence = re.sub(
        r"^(prosím|prosim)\s+",
        "",
        first_sentence,
        flags=re.IGNORECASE,
    ).strip()
    first_sentence = first_sentence.rstrip(" .,;:-")
    generic_titles = {"", "ahoj", "dobrý deň", "dobry den", "čau", "cau", "hello", "hi"}
    if first_sentence.lower() in generic_titles or len(first_sentence) < 4:
        return "New chat" if language == "en" else "Nový chat"
    words = first_sentence.split()
    normalized = " ".join(words[:7]).strip()
    normalized = normalized[0].upper() + normalized[1:]
    return normalized[:68].rstrip(" .,;:-")


def derive_thread_title(messages, language="sk", existing_title=""):
    generic = {"Nový chat", "New chat", "", None}
    if str(existing_title or "").strip() not in generic:
        return str(existing_title).strip()
    user_messages = []
    for message in messages or []:
        if str(message.get("role") or "") != "user":
            continue
        content = str(message.get("content") or "").strip()
        if content:
            user_messages.append(content)
    if not user_messages:
        return "New chat" if language == "en" else "Nový chat"
    candidate = build_thread_title(" ".join(user_messages[-4:]), language=language)
    if candidate not in generic:
        return candidate
    merged = " ".join(user_messages[-3:])
    candidate = build_thread_title(merged, language=language)
    return candidate if candidate not in generic else ("New chat" if language == "en" else "Nový chat")


def delete_assistant_thread(connection, user_id, thread_id):
    row = connection.execute(
        "SELECT * FROM assistant_threads WHERE id = ? AND user_id = ?",
        (thread_id, user_id),
    ).fetchone()
    if not row:
        return None
    connection.execute("DELETE FROM assistant_messages WHERE thread_id = ? AND user_id = ?", (thread_id, user_id))
    connection.execute("DELETE FROM assistant_threads WHERE id = ? AND user_id = ?", (thread_id, user_id))
    connection.commit()
    return row


def create_assistant_thread(connection, user_id, title="Nový chat"):
    now_value = format_timestamp(utc_now())
    cursor = connection.execute(
        """
        INSERT INTO assistant_threads (user_id, title, created_at, updated_at)
        VALUES (?, ?, ?, ?)
        """,
        (user_id, title or "Nový chat", now_value, now_value),
    )
    connection.commit()
    return cursor.lastrowid


def get_assistant_threads(connection, user_id, limit=40):
    rows = connection.execute(
        """
        SELECT
            assistant_threads.*,
            (
                SELECT COUNT(*)
                FROM assistant_messages
                WHERE assistant_messages.thread_id = assistant_threads.id
            ) AS message_count,
            (
                SELECT content
                FROM assistant_messages
                WHERE assistant_messages.thread_id = assistant_threads.id
                ORDER BY assistant_messages.id DESC
                LIMIT 1
            ) AS last_message,
            (
                SELECT created_at
                FROM assistant_messages
                WHERE assistant_messages.thread_id = assistant_threads.id
                ORDER BY assistant_messages.id DESC
                LIMIT 1
            ) AS last_message_at
        FROM assistant_threads
        WHERE user_id = ?
        ORDER BY updated_at DESC, id DESC
        LIMIT ?
        """,
        (user_id, limit),
    ).fetchall()
    return [
        {
            "id": row["id"],
            "title": row["title"] or "Nový chat",
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
            "message_count": row["message_count"] or 0,
            "last_message": row["last_message"] or "",
            "last_message_at": row["last_message_at"] or "",
        }
        for row in rows
    ]


def get_latest_assistant_thread(connection, user_id):
    row = connection.execute(
        """
        SELECT *
        FROM assistant_threads
        WHERE user_id = ?
        ORDER BY updated_at DESC, id DESC
        LIMIT 1
        """,
        (user_id,),
    ).fetchone()
    return row


def resolve_assistant_thread(connection, user_id, thread_id):
    if thread_id:
        row = connection.execute(
            "SELECT * FROM assistant_threads WHERE id = ? AND user_id = ?",
            (thread_id, user_id),
        ).fetchone()
        if row:
            return row
    return get_latest_assistant_thread(connection, user_id)


def get_assistant_messages(connection, user_id, thread_id=None, limit=24):
    resolved_thread = resolve_assistant_thread(connection, user_id, thread_id)
    if not resolved_thread:
        return []
    rows = connection.execute(
        """
        SELECT *
        FROM assistant_messages
        WHERE user_id = ? AND thread_id = ?
        ORDER BY id DESC
        LIMIT ?
        """,
        (user_id, resolved_thread["id"], limit),
    ).fetchall()
    messages = []
    for row in reversed(rows):
        messages.append(
            {
                "id": row["id"],
                "thread_id": row["thread_id"],
                "role": row["role"],
                "content": row["content"],
                "meta": safe_json_loads(row["meta_json"], {}),
                "review_status": row["review_status"] or "unreviewed",
                "reviewed_by": row["reviewed_by"] or 0,
                "reviewed_at": row["reviewed_at"] or "",
                "created_at": row["created_at"],
            }
        )
    return messages


def touch_assistant_thread(connection, thread_id, title=None):
    now_value = format_timestamp(utc_now())
    if title:
        connection.execute(
            "UPDATE assistant_threads SET title = ?, updated_at = ? WHERE id = ?",
            (title, now_value, thread_id),
        )
    else:
        connection.execute(
            "UPDATE assistant_threads SET updated_at = ? WHERE id = ?",
            (now_value, thread_id),
        )


def save_assistant_message(connection, user_id, thread_id, role, content, meta=None, language="sk"):
    review_status = "approved" if role == "user" else "unreviewed"
    meta_json = json.dumps(meta or {}, ensure_ascii=False)
    cursor = connection.execute(
        """
        INSERT INTO assistant_messages (user_id, thread_id, role, content, meta_json, review_status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (user_id, thread_id, role, content, meta_json, review_status, format_timestamp(utc_now())),
    )
    touch_assistant_thread(connection, thread_id)
    connection.commit()
    return cursor.lastrowid


def get_admin_assistant_thread_detail(connection, user_id, thread_id=None, limit=80):
    thread_row = resolve_assistant_thread(connection, user_id, thread_id)
    if not thread_row:
        return {"thread": None, "messages": []}
    return {
        "thread": {
            "id": thread_row["id"],
            "title": thread_row["title"] or "Nový chat",
            "created_at": thread_row["created_at"],
            "updated_at": thread_row["updated_at"],
        },
        "messages": get_assistant_messages(connection, user_id, thread_row["id"], limit=limit),
    }


def refresh_assistant_profile_memory(connection, user_id):
    rows = connection.execute(
        """
        SELECT content, created_at
        FROM assistant_messages
        WHERE user_id = ? AND role = 'user'
        ORDER BY id DESC
        LIMIT 12
        """,
        (user_id,),
    ).fetchall()
    recent_topics = []
    for row in rows:
        title = build_thread_title(row["content"] or "", language="sk")
        if title in {"Nový chat", "New chat", "Starší chat"}:
            continue
        if title not in recent_topics:
            recent_topics.append(title)
    focus_value = recent_topics[0] if recent_topics else ""
    notes_value = " | ".join(recent_topics[:5])
    now_value = format_timestamp(utc_now())
    connection.execute(
        """
        INSERT INTO assistant_profiles (user_id, focus, notes, updated_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            focus = excluded.focus,
            notes = excluded.notes,
            updated_at = excluded.updated_at
        """,
        (user_id, focus_value, notes_value, now_value),
    )
    connection.commit()


def get_assistant_usage_stats(connection, user_id):
    counts_row = connection.execute(
        """
        SELECT
            COUNT(DISTINCT thread_id) AS thread_count,
            COUNT(*) AS message_count,
            SUM(CASE WHEN role = 'assistant' THEN 1 ELSE 0 END) AS assistant_count,
            SUM(CASE WHEN role = 'assistant' AND instr(meta_json, '"used_web_search": true') > 0 THEN 1 ELSE 0 END) AS web_count,
            SUM(CASE WHEN instr(meta_json, 'attachment_preview') > 0 OR instr(meta_json, '"used_image": true') > 0 OR instr(meta_json, '"attachments"') > 0 THEN 1 ELSE 0 END) AS image_count,
            MAX(created_at) AS last_message_at
        FROM assistant_messages
        WHERE user_id = ?
        """,
        (user_id,),
    ).fetchone()
    profile = get_assistant_profile(connection, user_id)
    topics = [item.strip() for item in str(profile.get("notes") or "").split("|") if item.strip()]
    message_count = int(counts_row["message_count"] or 0)
    thread_count = int(counts_row["thread_count"] or 0)
    web_count = int(counts_row["web_count"] or 0)
    image_count = int(counts_row["image_count"] or 0)
    relevance = 0
    if message_count:
        relevance = min(96, 35 + min(thread_count, 6) * 7 + min(message_count, 30) // 2 + min(web_count, 8) * 3)
    return {
        "focus": profile.get("focus") or "",
        "topics": topics[:6],
        "updated_at": profile.get("updated_at") or "",
        "thread_count": thread_count,
        "message_count": message_count,
        "assistant_count": int(counts_row["assistant_count"] or 0),
        "web_count": web_count,
        "image_count": image_count,
        "last_message_at": counts_row["last_message_at"] or "",
        "relevance_percent": relevance,
    }


def get_activity_system_stats(connection):
    threshold = format_timestamp(utc_now() - timedelta(days=30))
    total_logs = int(connection.execute("SELECT COUNT(*) AS total FROM activity_logs").fetchone()["total"] or 0)
    recent_logs = int(
        connection.execute(
            "SELECT COUNT(*) AS total FROM activity_logs WHERE created_at >= ?",
            (threshold,),
        ).fetchone()["total"]
        or 0
    )
    unique_ips = set()
    admin_actions = 0
    rows = connection.execute(
        "SELECT event_type, event_meta FROM activity_logs ORDER BY created_at DESC LIMIT 250"
    ).fetchall()
    for row in rows:
        try:
            meta = json.loads(row["event_meta"] or "{}")
        except Exception:
            meta = {}
        ip_value = str(meta.get("ip") or "").strip()
        if ip_value:
            unique_ips.add(ip_value)
        if str(row["event_type"] or "").startswith("admin_"):
            admin_actions += 1
    return {
        "total_logs": total_logs,
        "recent_logs": recent_logs,
        "recent_unique_ips": len(unique_ips),
        "recent_admin_actions": admin_actions,
        "deleted_accounts": int(connection.execute("SELECT COUNT(*) AS total FROM activity_logs WHERE event_type = 'admin_account_deleted'").fetchone()["total"] or 0),
        "deactivated_accounts": int(connection.execute("SELECT COUNT(*) AS total FROM activity_logs WHERE event_type = 'admin_membership_deactivated'").fetchone()["total"] or 0),
    }


def extract_message_attachments(meta):
    meta = meta or {}
    attachments = meta.get("attachments") if isinstance(meta, dict) else None
    normalized = []
    if isinstance(attachments, list):
        for item in attachments:
            if not isinstance(item, dict):
                continue
            preview = str(item.get("attachment_preview") or "").strip()
            if not preview:
                continue
            normalized.append(
                {
                    "attachment_name": str(item.get("attachment_name") or "obrazok").strip() or "obrazok",
                    "attachment_type": str(item.get("attachment_type") or "").strip(),
                    "attachment_preview": preview,
                }
            )
    if normalized:
        return normalized
    attachment_preview = str(meta.get("attachment_preview") or "").strip() if isinstance(meta, dict) else ""
    if not attachment_preview:
        return []
    return [
        {
            "attachment_name": str(meta.get("attachment_name") or "obrazok").strip() or "obrazok",
            "attachment_type": str(meta.get("attachment_type") or "").strip(),
            "attachment_preview": attachment_preview,
        }
    ]


def build_openai_chat_messages(user_row, profile, history_messages, user_message, attachments=None, language="sk"):
    messages = [
        {
            "role": "system",
            "content": build_assistant_system_prompt(user_row, profile, language=language),
        }
    ]
    for message in history_messages[-10:]:
        role = "assistant" if message["role"] == "assistant" else "user"
        meta = message.get("meta") or {}
        history_attachments = extract_message_attachments(meta)
        if role == "user" and history_attachments:
            content_items = [{"type": "text", "text": message.get("content") or "Pozri priložené obrázky v predchádzajúcej správe."}]
            for item in history_attachments:
                content_items.append({"type": "image_url", "image_url": {"url": item["attachment_preview"]}})
            messages.append(
                {
                    "role": "user",
                    "content": content_items,
                }
            )
        else:
            messages.append(
                {
                    "role": role,
                    "content": message.get("content") or "",
                }
            )
    if attachments:
        content_items = [{"type": "text", "text": user_message or "Vyhodnoť priložené obrázky a poraď mi ďalší krok."}]
        for item in attachments:
            content_items.append({"type": "image_url", "image_url": {"url": item["data_url"]}})
        messages.append(
            {
                "role": "user",
                "content": content_items,
            }
        )
    else:
        messages.append({"role": "user", "content": user_message})
    return messages


def extract_openai_response_text(payload):
    output_text = str(payload.get("output_text") or "").strip()
    if output_text:
        return output_text

    choices = payload.get("choices") or []
    if choices:
        message = choices[0].get("message") or {}
        content = message.get("content") or ""
        if isinstance(content, str) and content.strip():
            return content.strip()
        if isinstance(content, list):
            chunks = []
            for item in content:
                if isinstance(item, str) and item.strip():
                    chunks.append(item.strip())
                    continue
                if not isinstance(item, dict):
                    continue
                if item.get("type") in {"text", "output_text"}:
                    text_value = item.get("text") or item.get("value") or ""
                    if isinstance(text_value, dict):
                        text_value = text_value.get("value", "")
                    if text_value:
                        chunks.append(str(text_value).strip())
            if chunks:
                return "\n\n".join(chunks).strip()

    chunks = []
    for item in payload.get("output", []) or []:
        if isinstance(item, dict) and item.get("type") in {"output_text", "text"}:
            text_value = item.get("text") or item.get("value") or ""
            if isinstance(text_value, dict):
                text_value = text_value.get("value", "")
            if text_value:
                chunks.append(str(text_value).strip())
        for content in item.get("content", []) or []:
            content_type = content.get("type", "")
            if content_type in {"output_text", "text"}:
                text_value = content.get("text") or content.get("value") or ""
                if isinstance(text_value, dict):
                    text_value = text_value.get("value", "")
                if text_value:
                    chunks.append(str(text_value).strip())
    return "\n\n".join(chunk for chunk in chunks if chunk).strip()


def perform_openai_request(url, payload):
    request = urllib_request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib_request.urlopen(request, timeout=90) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        try:
            detail = exc.read().decode("utf-8")
        except Exception:
            detail = ""
        raise RuntimeError(f"OpenAI API chyba: {detail or exc.reason}") from exc
    except urllib_error.URLError as exc:
        raise RuntimeError(f"OpenAI API nie je dostupné: {exc.reason}") from exc


def build_image_data_url(content_type, content_bytes):
    encoded = base64.b64encode(content_bytes).decode("ascii")
    return f"data:{content_type};base64,{encoded}"


def normalize_attachment_name(value):
    return " ".join(str(value or "").strip().split())[:120].strip() or "obrazok"


def parse_assistant_attachments(form):
    attachment_entries = []
    for key in ("attachment", "image", "file"):
        attachment_entries.extend(form.get(key) or [])
    attachments = []
    for attachment_entry in attachment_entries:
        if not attachment_entry or not attachment_entry.get("content"):
            continue
        content_type = str(attachment_entry.get("content_type") or "").strip().lower()
        if not content_type.startswith("image/"):
            raise ValueError("AI aktuálne podporuje iba obrázky.")
        content_bytes = attachment_entry.get("content") or b""
        if len(content_bytes) > MAX_AI_IMAGE_BYTES:
            raise ValueError(f"Obrázok je príliš veľký. Maximum je {format_mb(MAX_AI_IMAGE_BYTES)}.")
        filename = normalize_attachment_name(attachment_entry.get("filename") or "obrazok")
        attachments.append(
            {
                "filename": filename,
                "content_type": content_type,
                "content_bytes": content_bytes,
                "data_url": build_image_data_url(content_type, content_bytes),
            }
        )
    return attachments


def create_assistant_upload_session(connection, user_id):
    token = secrets.token_urlsafe(24)
    created_at = utc_now()
    expires_at = created_at + timedelta(minutes=10)
    connection.execute(
        """
        INSERT INTO assistant_upload_sessions (
            user_id, token, status, attachment_name, attachment_type, attachment_base64, created_at, expires_at, uploaded_at
        ) VALUES (?, ?, 'pending', '', '', '', ?, ?, '')
        """,
        (user_id, token, format_timestamp(created_at), format_timestamp(expires_at)),
    )
    connection.commit()
    return {
        "token": token,
        "created_at": format_timestamp(created_at),
        "expires_at": format_timestamp(expires_at),
    }


def get_assistant_upload_session(connection, token):
    if not token:
        return None
    row = connection.execute(
        "SELECT * FROM assistant_upload_sessions WHERE token = ?",
        (str(token).strip(),),
    ).fetchone()
    if not row:
        return None
    expires_at = parse_timestamp(row["expires_at"])
    if not expires_at or expires_at < utc_now():
        return None
    return row


def save_assistant_upload_to_session(connection, token, attachment):
    if not attachment:
        raise ValueError("Chýba obrázok na nahratie.")
    encoded = base64.b64encode(attachment["content_bytes"]).decode("ascii")
    now_value = format_timestamp(utc_now())
    connection.execute(
        """
        UPDATE assistant_upload_sessions
        SET status = 'uploaded',
            attachment_name = ?,
            attachment_type = ?,
            attachment_base64 = ?,
            uploaded_at = ?
        WHERE token = ?
        """,
        (attachment["filename"], attachment["content_type"], encoded, now_value, token),
    )
    connection.commit()


def build_assistant_upload_payload(row):
    if not row:
        return {"status": "missing"}
    if row["status"] != "uploaded" or not row["attachment_base64"]:
        return {
            "status": row["status"] or "pending",
            "uploaded": False,
            "expires_at": row["expires_at"] or "",
        }
    content_type = row["attachment_type"] or "image/jpeg"
    data_url = f"data:{content_type};base64,{row['attachment_base64']}"
    return {
        "status": "uploaded",
        "uploaded": True,
        "attachment_name": row["attachment_name"] or "obrazok",
        "attachment_type": content_type,
        "attachment_preview": data_url,
        "expires_at": row["expires_at"] or "",
        "uploaded_at": row["uploaded_at"] or "",
    }


def build_assistant_system_prompt(user_row, profile, language="sk"):
    now_value = datetime.now().astimezone()
    today_value = now_value.strftime("%d.%m.%Y")
    time_value = now_value.strftime("%H:%M")
    brief = (profile or {}).get("brief") or {}
    counts = brief.get("counts") or {}
    focus = brief.get("focus") or []
    suggestions = brief.get("suggestions") or []
    focus_line = " | ".join(str(item).strip() for item in focus[:2] if str(item).strip())
    suggestion_line = " | ".join(str(item.get("title") or "").strip() for item in suggestions[:3] if str(item.get("title") or "").strip())
    stored_focus = str((profile or {}).get("focus") or "").strip()
    brief_summary_sk = (
        f"Interný denný kontext: otvorené úlohy {counts.get('open_tasks', 0)}, "
        f"dnes {counts.get('due_today', 0)}, omeškané {counts.get('overdue', 0)}, "
        f"vysoká priorita {counts.get('high_priority', 0)}."
    )
    brief_summary_en = (
        f"Internal day context: open tasks {counts.get('open_tasks', 0)}, "
        f"due today {counts.get('due_today', 0)}, overdue {counts.get('overdue', 0)}, "
        f"high priority {counts.get('high_priority', 0)}."
    )
    brief_append_sk = brief_summary_sk + " "
    if focus_line:
        brief_append_sk += f"Aktuálne fokusy dňa: {focus_line}. "
    if suggestion_line:
        brief_append_sk += f"Navrhované pracovné okruhy: {suggestion_line}. "
    if stored_focus:
        brief_append_sk += f"Uložený fokus používateľa: {stored_focus}. "

    brief_append_en = brief_summary_en + " "
    if focus_line:
        brief_append_en += f"Current focus cues: {focus_line}. "
    if suggestion_line:
        brief_append_en += f"Suggested work tracks: {suggestion_line}. "
    if stored_focus:
        brief_append_en += f"Stored user focus: {stored_focus}. "

    if language == "en":
        return (
            "You are Unifyo AI, a professional internal assistant for financial intermediaries working in Slovakia. "
            "Help with daily organisation, priorities, follow-ups, client communication, meeting prep, call scripts, emails, SMS, pipeline and operations. "
            "Always communicate in natural, concise, professional English, while keeping Slovak market context in mind. "
            "Internally evaluate: 1) user intent, 2) question type, 3) Slovak market context, 4) best response format. Never show this hidden process. "
            "Default response structure: 1) direct answer, 2) short explanation, 3) recommended next step if useful. "
            "Do not print explicit section labels such as Direct answer, Explanation or Next step. Present the answer as one clean, natural response. "
            "Keep the first version concise and easy to use. People prefer less text. If more detail would help, end with a short sentence offering an extended version on request. "
            "Do not add sources automatically. Mention them only when genuinely useful, especially for regulatory or official topics. "
            "If a source is important, add only one minimalist final line in this exact style: Source: https://... "
            "Prefer official sources such as nbs.sk, slov-lex.sk, official financial institution sites and, when relevant, alanrampacek.sk or prosight.sk. "
            f"The current reference date and time are {today_value} {time_value}. If the user asks for the current date or time, answer directly from this reference. "
            "For all allowed in-domain topics, first use current public verification when technically available before composing the answer. "
            "When web context is used, internally compare multiple relevant public sources when available and prefer the answer only after cross-checking at least two independent sources. "
            "For current bank names, mortgage providers, interest rates, product availability, market changes and similar current market facts, do not rely on memory. Answer only if you have current public verification; otherwise clearly say you cannot verify the latest state right now. "
            "Do not say that you have no access to web search or that browsing is impossible if web verification is available in the current workflow. If the latest public information could not be confirmed, say only that the current public state could not be reliably verified right now. "
            "Your default mode is work guidance for the user's day. Prefer helping with tasks, clients, follow-ups, priorities, objections, emails, meetings and next actions. "
            "You are only allowed to meaningfully engage in topics related to financial intermediation, finance, banking, insurance, investments, mortgages, client communication, sales, business workflow, regulation, economics and closely related work topics. "
            "You must not meaningfully answer questions about unrelated topics such as medicine, diseases, health symptoms, relationships, entertainment, sport, hobbies, general trivia or other off-topic areas. "
            "If the user asks about an unrelated topic, do not provide a substantive answer to that topic, do not explore it, and do not improvise. Briefly apologize, state that Unifyo AI is specialized in financial intermediation and related work topics, and immediately redirect to a relevant work question such as: How can I help with your clients, priorities or financial communication today? "
            "For unrelated topics, keep the answer very short, ideally within one or two sentences plus the redirect. "
            "When possible, anchor the answer in action: what to send, whom to call, what to prioritise, or what the next best step is. "
            "If the user asks for contact cleanup or file compression, briefly explain that the actual processing is available in the app module and naturally suggest opening the relevant section of the Unifyo app. "
            "Generate a table only when the user explicitly asks for a table or a side-by-side comparison. Structure it as a clean comparison table or compact structured rows, never markdown pipe syntax. "
            "If the user asks for a PDF, prepare only the PDF-ready content or structure that can be exported in the app. Never claim that a file was generated, attached or downloaded unless the system really provided it. "
            "If the user asks for text, first provide the shortest useful version. Then offer an expanded version only on request. "
            "End many answers with one practical work-oriented question or suggestion, but do not sound repetitive or robotic. "
            "Never invent facts, names, regulations, rates, product details, legal interpretations or source claims. "
            "Only state something as factual if it is supported by the provided context, verified history, image content, or reliable public information available to you. "
            "If you are not sure, say that clearly and naturally. Distinguish between verified information, practical suggestion and assumption. "
            "If a claim is not verified, do not present it as true. "
            "Do not provide binding legal, tax or regulated investment advice; suggest checking with compliance or a qualified expert. "
            "If the user attaches an image or screenshot, first evaluate it clearly and then suggest a next step. "
            "Do not use markdown asterisks or noisy formatting. Write in short paragraphs or simple bullets without markdown artefacts. "
            "If helpful, proactively offer a short email/SMS version or a concrete next step. "
            f"{brief_append_en}"
            f"The user's name is {user_row['name'] or user_row['email']}. "
            f"Use internal prompt version {PROMPT_VERSION}, but do not mention it unless necessary."
        )
    return (
        "Si Unifyo AI, profesionálny interný asistent pre finančných sprostredkovateľov na Slovensku. "
        "Pomáhaš s organizáciou dňa, prioritami, follow-upmi, klientskou komunikáciou, prípravou stretnutí, "
        "call scriptami, e-mailami, SMS správami, pipeline a každodennou operatívou. "
        "Komunikuj vždy po slovensky, prirodzene, vecne, profesionálne a prakticky. "
        "Najprv si interne (skryto) vyhodnoť otázku v krokoch: "
        "1) zámer používateľa, 2) typ otázky (praktická/odborná/obchodná/informačná), "
        "3) kontext slovenského trhu a použiteľnosť, 4) najlepší formát odpovede. "
        "Tento interný postup nikdy nevypisuj. Používateľovi ukáž iba finálnu odpoveď. "
        "Predvolený výstup drž v 3 častiach: "
        "1) Priama odpoveď, 2) Stručné vysvetlenie, 3) Odporúčaný ďalší krok (ak je vhodný). "
        "Tieto časti však neoznačuj doslovnými nadpismi ako Priama odpoveď, Stručné vysvetlenie alebo Odporúčaný ďalší krok. Odpoveď má pôsobiť ako jeden prirodzený, čistý text. "
        "Prvú verziu odpovede drž stručnú a výstižnú, lebo ľudia nemajú radi veľa textu. "
        "Ak je vhodné doplniť viac detailov, ukonči odpoveď krátkou vetou, že môžeš doplniť rozšírenú verziu na požiadanie. "
        "Zdroje alebo odkazy neuvádzaj automaticky. Uveď ich len vtedy, keď sú skutočne potrebné "
        "alebo užitočné (najmä pri regulačných, právnych, dohľadových a oficiálnych témach). "
        "Ak je dôležité uviesť zdroj, daj na úplný koniec iba jeden minimalistický riadok v tvare: Zdroj: https://... "
        "Ak je vhodné odporučiť zdroj, preferuj oficiálne a dôveryhodné weby: nbs.sk, slov-lex.sk, "
        "oficiálne stránky finančných inštitúcií, prípadne podľa kontextu alanrampacek.sk a prosight.sk. "
        f"Aktuálny referenčný dátum a čas sú {today_value} {time_value}. Pri časovo citlivých témach (novinky, legislatíva, sadzby, zmeny pravidiel) "
        "Ak sa používateľ pýta priamo na aktuálny dátum alebo čas, odpovedz priamo z tohto referenčného dátumu a času a netvrď, že k nim nemáš prístup. "
        "pri všetkých povolených témach najprv pracuj s aktuálnym verejným overením, ak je technicky dostupné. "
        "Ak používaš webový kontext, vnútorne porovnaj viac relevantných verejných zdrojov, keď sú dostupné, a preferuj odpoveď až po overení aspoň z dvoch nezávislých zdrojov. "
        "Pri aktuálnych názvoch bánk, poskytovateľoch hypoték, úrokových sadzbách, dostupnosti produktov a podobných aktuálnych trhových faktoch sa nespoliehaj na pamäť. Odpovedaj len vtedy, keď máš aktuálne verejné overenie; inak jasne povedz, že najnovší stav nevieš potvrdiť. "
        "Nehovor, že nemáš prístup k vyhľadávaniu alebo že web nevieš použiť, ak je webové overenie v aktuálnom workflow dostupné. Ak sa najnovší verejný stav nepodarí potvrdiť, povedz len to, že aktuálny verejný stav sa teraz nepodarilo spoľahlivo overiť. "
        "Tvoj predvolený režim je pomoc s pracovným dňom používateľa. Prirodzene smeruj odpovede k úlohám, klientom, follow-upom, prioritám, argumentácii, komunikácii a ďalšiemu kroku. "
        "Vecne sa venuj len témam súvisiacim s finančným sprostredkovaním, financiami, bankami, poistením, investíciami, hypotékami, klientskou komunikáciou, obchodom, reguláciou, ekonomikou a blízkou pracovnou praxou. "
        "Nesmieš vecne riešiť nesúvisiace témy ako choroby, zdravotné ťažkosti, medicínu, vzťahy, šport, zábavu, hobby, všeobecné vedomosti ani iné oblasti mimo tejto domény. "
        "Ak sa používateľ spýta na inú tému, túto tému vôbec nerozvíjaj, neodpovedaj na ňu odborne ani orientačne a nič si k nej nedomýšľaj. Stručne sa ospravedlň, povedz, že Unifyo AI je zameraná na finančné sprostredkovanie a súvisiacu prácu, a potom používateľa presmeruj na relevantnú otázku typu: Ako ti viem dnes pomôcť s klientmi, úlohami, prioritami alebo finančnou komunikáciou? "
        "Pri nesúvisiacich témach drž odpoveď veľmi krátku, ideálne v jednej až dvoch vetách plus presmerovanie. "
        "Keď sa dá, odpoveď ukotvi do akcie: čo poslať, komu zavolať, čo vysvetliť klientovi, čo prioritizovať alebo aký je ďalší najlepší krok. "
        "Ak používateľ potrebuje čistenie kontaktov alebo kompresiu súborov, stručne vysvetli, že samotné spracovanie je dostupné v príslušnej časti aplikácie Unifyo, a prirodzene navrhni otvorenie správneho modulu. "
        "Tabuľku generuj len vtedy, keď si ju používateľ výslovne pýta alebo keď pýta priame porovnanie viacerých možností. Priprav ju ako čistú porovnávaciu tabuľku alebo kompaktné štruktúrované riadky, nie v markdown pipe zápise. "
        "Ak používateľ pýta PDF, priprav len obsah alebo osnovu vhodnú na reálny export do PDF v aplikácii. Nikdy netvrď, že súbor bol vytvorený, priložený alebo stiahnuteľný, pokiaľ ho systém reálne nepridal. "
        "Pri požiadavke na text najprv daj krátku použiteľnú verziu. Rozšírenú verziu ponúkni len na požiadanie. "
        "Veľa odpovedí ukonči jednou stručnou pracovnou otázkou alebo návrhom ďalšieho kroku, ale neopakuj sa mechanicky v každej odpovedi. "
        "Nikdy si nevymýšľaj fakty, sadzby, mená, pravidlá, legislatívu, produktové parametre ani zdroje. "
        "Ako fakt uvádzaj len to, čo máš overené z kontextu, histórie chatu, priloženého obrázka alebo dôveryhodných verejných informácií, ak sú dostupné. "
        "Ak si niečím nie si istý, povedz to jasne a prirodzene. Rozlišuj medzi overenou informáciou, praktickým odporúčaním a odhadom. "
        "Čokoľvek neoverené nesmieš prezentovať ako istotu. "
        "Nedávaj záväzné právne, daňové ani regulované investičné odporúčania; pri takých otázkach "
        "odporuč overenie cez compliance alebo kvalifikovaného odborníka. "
        "Ak používateľ priloží obrázok alebo screenshot, najprv ho vecne vyhodnoť a potom odporuč konkrétny ďalší krok. "
        "Nepoužívaj markdown hviezdičky ani zbytočné formátovanie typu **text**. "
        "Píš čisto, priamo a štruktúrovane pomocou krátkych odsekov alebo jednoduchých bodov bez markdown artefaktov. "
        "Ak to pomôže používateľovi, proaktívne ponúkni krátku verziu do e-mailu/SMS alebo konkrétny ďalší krok. "
        f"{brief_append_sk}"
        f"Používateľ sa volá {user_row['name'] or user_row['email']}. "
        f"Používaj internú prompt verziu {PROMPT_VERSION}, ale túto informáciu bežne nevypisuj používateľovi."
    )


def should_use_openai_web_search(user_message):
    if not OPENAI_WEB_SEARCH_ENABLED or OPENAI_WEB_SEARCH_RUNTIME_DISABLED:
        return False
    text = str(user_message or "").strip().lower()
    if not text:
        return False
    return True


def requires_verified_current_info(user_message):
    text = str(user_message or "").strip().lower()
    if not text:
        return False
    strict_markers = (
        "banka",
        "banky",
        "bank",
        "banks",
        "hypo",
        "hypot",
        "hypotek",
        "mortgage",
        "mortgages",
        "úrok",
        "urok",
        "interest",
        "sadzb",
        "rate",
        "rates",
        "refinanc",
        "fixáci",
        "fixac",
        "úver",
        "uver",
        "loan",
        "loans",
        "aktuálne",
        "aktualne",
        "current",
        "latest",
        "dnes",
        "today",
    )
    return any(marker in text for marker in strict_markers)


def call_openai_assistant(user_row, profile, history_messages, user_message, attachments=None, language="sk"):
    global OPENAI_WEB_SEARCH_RUNTIME_DISABLED
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY nie je nastavený, takže AI asistent zatiaľ nemôže odpovedať.")

    response_meta = {
        "prompt_version": PROMPT_VERSION,
        "model": OPENAI_MODEL,
        "used_web_search": False,
        "used_image": bool(attachments),
        "attachment_name": attachments[0]["filename"] if attachments else "",
        "attachment_count": len(attachments or []),
    }
    history_has_images = any(
        bool(extract_message_attachments(message.get("meta") or {}))
        for message in history_messages[-10:]
        if message.get("role") == "user"
    )
    if history_has_images:
        response_meta["used_image"] = True

    if attachments or history_has_images:
        fallback_messages = build_openai_chat_messages(
            user_row,
            profile,
            history_messages,
            user_message,
            attachments=attachments,
            language=language,
        )
        fallback_payload = perform_openai_request(
            "https://api.openai.com/v1/chat/completions",
            {
                "model": "gpt-4.1-mini",
                "messages": fallback_messages,
                "temperature": 0.5,
                "max_tokens": 700,
            },
        )
        text = extract_openai_response_text(fallback_payload)
        if text:
            response_meta["model"] = "gpt-4.1-mini"
            return text, response_meta
        raise RuntimeError("AI asistent vrátil prázdnu odpoveď k priloženému obrázku.")

    input_messages = []
    for message in history_messages[-10:]:
        input_messages.append(
            {
                "role": "assistant" if message["role"] == "assistant" else "user",
                "content": message["content"],
            }
        )
    input_messages.append(
        {
            "role": "user",
            "content": user_message,
        }
    )

    payload_base = {
        "model": OPENAI_MODEL,
        "instructions": build_assistant_system_prompt(user_row, profile, language=language),
        "input": input_messages,
        "max_output_tokens": 500,
    }
    primary_error = ""
    response_attempts = []
    use_web_search = should_use_openai_web_search(user_message)
    strict_current_info = requires_verified_current_info(user_message)
    if use_web_search:
        response_attempts.append({**payload_base, "tools": [{"type": OPENAI_WEB_SEARCH_TOOL}]})
    response_attempts.append(payload_base)

    for candidate_payload in response_attempts:
        try:
            primary_payload = perform_openai_request("https://api.openai.com/v1/responses", candidate_payload)
            text = extract_openai_response_text(primary_payload)
            if text:
                response_meta["used_web_search"] = bool("tools" in candidate_payload)
                response_meta["model"] = candidate_payload.get("model", OPENAI_MODEL)
                if response_meta["used_web_search"]:
                    source_urls = extract_response_source_urls(primary_payload)
                    if not source_urls:
                        source_urls = extract_urls_from_text(text)
                    source_count = len({
                        (urlparse(url).hostname or "").lower().removeprefix("www.")
                        for url in source_urls
                        if (urlparse(url).hostname or "").strip()
                    })
                    response_meta["web_sources"] = source_urls[:4]
                    response_meta["web_source_count"] = source_count
                    response_meta["web_confidence_percent"] = estimate_web_confidence(source_urls)
                    if strict_current_info and source_count < 1:
                        primary_error = (
                            "Aktuálny stav sa nepodarilo spoľahlivo overiť z verejných zdrojov, "
                            "preto AI radšej nevráti neoverené údaje."
                        )
                        continue
                return text, response_meta
            primary_error = f"AI asistent vrátil prázdnu odpoveď. Stav: {primary_payload.get('status') or 'unknown'}."
        except RuntimeError as exc:
            primary_error = str(exc)
            if use_web_search and "tools" in candidate_payload:
                lowered = primary_error.lower()
                if (
                    "tool" in lowered
                    and ("unknown" in lowered or "unsupported" in lowered or "invalid" in lowered)
                ) or OPENAI_WEB_SEARCH_TOOL.lower() in lowered:
                    OPENAI_WEB_SEARCH_RUNTIME_DISABLED = True

    if strict_current_info:
        if language == "en":
            raise RuntimeError(
                "I can’t reliably confirm the latest market data from current public sources right now, "
                "so I’d rather not give you potentially outdated information. Try asking again in a moment or name the specific bank or product you want to verify."
            )
        raise RuntimeError(
            "Aktuálne trhové údaje sa mi teraz nepodarilo spoľahlivo overiť z verejných zdrojov, "
            "preto ti radšej nedám potenciálne zastaranú informáciu. Skús otázku zopakovať o chvíľu alebo uveď konkrétnu banku či produkt, ktorý chceš overiť."
        )

    fallback_messages = [
        {
            "role": "system",
            "content": build_assistant_system_prompt(user_row, profile, language=language),
        }
    ]
    for message in history_messages[-8:]:
        fallback_messages.append(
            {
                "role": "assistant" if message["role"] == "assistant" else "user",
                "content": message["content"],
            }
        )
    fallback_messages.append({"role": "user", "content": user_message})

    fallback_payload = perform_openai_request(
        "https://api.openai.com/v1/chat/completions",
        {
            "model": "gpt-4.1-mini",
            "messages": fallback_messages,
            "temperature": 0.7,
            "max_tokens": 500,
        },
    )
    text = extract_openai_response_text(fallback_payload)
    if text:
        response_meta["model"] = "gpt-4.1-mini"
        return text, response_meta
    raise RuntimeError(primary_error or "AI asistent vrátil prázdnu odpoveď.")


def is_membership_active(membership):
    if not membership:
        return False
    if membership.get("status") not in {"active", "trialing"}:
        return False
    period_end = membership.get("current_period_end")
    if not period_end:
        return False
    return period_end > utc_now()


def user_has_service_access(user_row, membership=None):
    if not user_row:
        return False
    if user_row["role"] == "admin":
        return True
    return is_membership_active(membership)


def get_membership_display_state(membership):
    active = is_membership_active(membership)
    if not membership:
        return {
            "active": False,
            "status": "inactive",
            "valid_until": "",
            "started_at": "",
            "next_renewal_at": "",
            "cancelled_at": "",
            "cancel_at_period_end": False,
        }
    if not active:
        return {
            "active": False,
            "status": "inactive",
            "valid_until": "",
            "started_at": "",
            "next_renewal_at": "",
            "cancelled_at": format_timestamp(membership.get("cancelled_at")) if membership.get("cancelled_at") else "",
            "cancel_at_period_end": bool(membership.get("cancel_at_period_end")),
        }
    return {
        "active": True,
        "status": membership.get("status") or "active",
        "valid_until": format_timestamp(membership.get("current_period_end")) if membership.get("current_period_end") else "",
        "started_at": format_timestamp(membership.get("created_at")) if membership.get("created_at") else "",
        "next_renewal_at": format_timestamp(membership.get("next_renewal_at")) if membership.get("next_renewal_at") else "",
        "cancelled_at": format_timestamp(membership.get("cancelled_at")) if membership.get("cancelled_at") else "",
        "cancel_at_period_end": bool(membership.get("cancel_at_period_end")),
    }


def user_payload(connection, user_row, headers=None):
    membership = get_membership(connection, user_row["id"])
    display_membership = get_membership_display_state(membership)
    registration_consent = get_latest_registration_consent(connection, user_row["id"])
    checkout_consent = get_latest_checkout_consent(connection, user_row["id"])
    return {
        "authenticated": True,
        "user": {
            "id": user_row["id"],
            "name": user_row["name"],
            "email": user_row["email"],
            "role": user_row["role"],
            "is_admin": user_row["role"] == "admin",
            "can_manage_admin_tools": can_manage_admin_tools(user_row),
            "created_at": user_row["created_at"],
            "membership_active": display_membership["active"],
            "membership_status": display_membership["status"],
            "membership_stripe_status": membership.get("stripe_status") if membership else "",
            "membership_valid_until": display_membership["valid_until"],
            "membership_started_at": display_membership["started_at"],
            "membership_next_renewal_at": display_membership["next_renewal_at"],
            "membership_cancelled_at": display_membership["cancelled_at"],
            "membership_cancel_at_period_end": display_membership["cancel_at_period_end"],
            "membership_internal_subscription_number": membership.get("internal_subscription_number") if membership else "",
            "membership_last_order_number": membership.get("last_order_number") if membership else "",
            "membership_last_checkout_session_id": membership.get("last_checkout_session_id") if membership else "",
            "membership_last_payment_intent_id": membership.get("last_payment_intent_id") if membership else "",
            "membership_last_invoice_id": membership.get("last_invoice_id") if membership else "",
            "membership_stripe_subscription_id": membership.get("stripe_subscription_id") if membership else "",
            "registration_consent_at": registration_consent.get("created_at") if registration_consent else "",
            "registration_marketing_consent": bool(registration_consent.get("marketing_consent")) if registration_consent else False,
            "checkout_consent_at": checkout_consent.get("created_at") if checkout_consent else "",
            "session_expires_at": get_session_expiry(connection, headers) if headers else "",
        },
    }


def get_request_base_url(handler):
    configured = str(APP_BASE_URL or "").strip()
    if configured:
        # Accept both full URLs and host-like values in env (e.g. unifyo.online).
        parsed = urlparse(configured if "://" in configured else f"https://{configured}")
        if parsed.scheme in {"http", "https"} and parsed.netloc:
            return f"{parsed.scheme}://{parsed.netloc}".rstrip("/")
    forwarded_proto = handler.headers.get("X-Forwarded-Proto")
    proto = forwarded_proto or ("https" if PORT == 443 else "http")
    host = handler.headers.get("X-Forwarded-Host") or handler.headers.get("Host") or f"127.0.0.1:{PORT}"
    if host:
        return f"{proto}://{host}".rstrip("/")
    return f"http://127.0.0.1:{PORT}"


def normalize_host_name(value):
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    first = raw.split(",", 1)[0].strip()
    parsed = urlparse(f"//{first}")
    hostname = (parsed.hostname or "").strip().lower()
    return hostname or first.split(":", 1)[0].strip().lower()


def get_or_create_customer(connection, user_row):
    customer_id = str(user_row["stripe_customer_id"] or "").strip()
    if customer_id and re.fullmatch(r"cus_[A-Za-z0-9]+", customer_id):
        return customer_id
    customer_id = ""

    customer = stripe.Customer.create(
        email=user_row["email"],
        metadata={"user_id": str(user_row["id"])},
    )
    customer_id = customer["id"]
    now = format_timestamp(utc_now())
    connection.execute(
        "UPDATE users SET stripe_customer_id = ?, updated_at = ? WHERE id = ?",
        (customer_id, now, user_row["id"]),
    )
    connection.commit()
    return customer_id


def sync_membership_from_subscription(connection, subscription, fallback_user_id=None):
    customer_id = subscription.get("customer")
    subscription_id = subscription.get("id")
    status = subscription.get("status", "inactive")
    period_end_ts = subscription.get("current_period_end")
    period_end = datetime.fromtimestamp(period_end_ts, tz=timezone.utc) if period_end_ts else None
    next_renewal_at = period_end
    cancel_at_period_end = bool(subscription.get("cancel_at_period_end"))
    cancelled_ts = subscription.get("canceled_at") or subscription.get("cancel_at")
    cancelled_at = datetime.fromtimestamp(cancelled_ts, tz=timezone.utc) if cancelled_ts else None
    latest_invoice_id = subscription.get("latest_invoice") or ""
    now = format_timestamp(utc_now())

    user_row = None
    if customer_id:
        user_row = connection.execute(
            "SELECT * FROM users WHERE stripe_customer_id = ?",
            (customer_id,),
        ).fetchone()
    if not user_row and customer_id:
        try:
            customer = stripe.Customer.retrieve(customer_id)
            customer_email = normalize_email(customer.get("email"))
        except Exception:
            customer_email = ""
        if customer_email:
            user_row = connection.execute(
                "SELECT * FROM users WHERE email = ?",
                (customer_email,),
            ).fetchone()
    if not user_row and fallback_user_id:
        user_row = connection.execute(
            "SELECT * FROM users WHERE id = ?",
            (fallback_user_id,),
        ).fetchone()
    if user_row and customer_id and (user_row["stripe_customer_id"] or "") != customer_id:
        connection.execute(
            "UPDATE users SET stripe_customer_id = ?, updated_at = ? WHERE id = ?",
            (customer_id, now, user_row["id"]),
        )
        user_row = connection.execute(
            "SELECT * FROM users WHERE id = ?",
            (user_row["id"],),
        ).fetchone()

    if not user_row:
        return

    existing_membership = get_membership(connection, user_row["id"])
    internal_subscription_number = (existing_membership or {}).get("internal_subscription_number") or generate_internal_subscription_number()
    last_order_number = (existing_membership or {}).get("last_order_number") or ""
    last_checkout_session_id = (existing_membership or {}).get("last_checkout_session_id") or ""
    last_payment_intent_id = (existing_membership or {}).get("last_payment_intent_id") or ""
    last_invoice_id = latest_invoice_id or ((existing_membership or {}).get("last_invoice_id") or "")

    connection.execute(
        """
        INSERT INTO memberships (
            user_id, status, stripe_customer_id, stripe_subscription_id, current_period_end, created_at, updated_at,
            stripe_status, cancel_at_period_end, cancelled_at, next_renewal_at,
            internal_subscription_number, last_order_number, last_checkout_session_id, last_payment_intent_id, last_invoice_id
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            status = excluded.status,
            stripe_customer_id = excluded.stripe_customer_id,
            stripe_subscription_id = excluded.stripe_subscription_id,
            current_period_end = excluded.current_period_end,
            stripe_status = excluded.stripe_status,
            cancel_at_period_end = excluded.cancel_at_period_end,
            cancelled_at = excluded.cancelled_at,
            next_renewal_at = excluded.next_renewal_at,
            internal_subscription_number = excluded.internal_subscription_number,
            last_order_number = CASE WHEN excluded.last_order_number != '' THEN excluded.last_order_number ELSE memberships.last_order_number END,
            last_checkout_session_id = CASE WHEN excluded.last_checkout_session_id != '' THEN excluded.last_checkout_session_id ELSE memberships.last_checkout_session_id END,
            last_payment_intent_id = CASE WHEN excluded.last_payment_intent_id != '' THEN excluded.last_payment_intent_id ELSE memberships.last_payment_intent_id END,
            last_invoice_id = CASE WHEN excluded.last_invoice_id != '' THEN excluded.last_invoice_id ELSE memberships.last_invoice_id END,
            updated_at = excluded.updated_at
        """,
        (
            user_row["id"],
            status,
            customer_id,
            subscription_id,
            format_timestamp(period_end) if period_end else "",
            now,
            now,
            status,
            1 if cancel_at_period_end else 0,
            format_timestamp(cancelled_at) if cancelled_at else "",
            format_timestamp(next_renewal_at) if next_renewal_at else "",
            internal_subscription_number,
            last_order_number,
            last_checkout_session_id,
            last_payment_intent_id,
            last_invoice_id,
        ),
    )
    record_subscription_event(
        connection,
        user_row["id"],
        "stripe_sync",
        status,
        {
            "stripe_subscription_id": subscription_id or "",
            "stripe_status": status,
            "current_period_end": format_timestamp(period_end) if period_end else "",
            "cancel_at_period_end": cancel_at_period_end,
            "cancelled_at": format_timestamp(cancelled_at) if cancelled_at else "",
            "last_invoice_id": last_invoice_id,
        },
    )
    log_activity(
        connection,
        "membership_updated",
        "Členstvo bolo synchronizované zo Stripe.",
        user_row["id"],
        status=status,
        current_period_end=format_timestamp(period_end) if period_end else "",
    )
    connection.commit()


def should_sync_membership_now(user_id, force=False):
    if force:
        LAST_MEMBERSHIP_SYNC_BY_USER[user_id] = utc_now()
        return True
    last = LAST_MEMBERSHIP_SYNC_BY_USER.get(user_id)
    if not last:
        LAST_MEMBERSHIP_SYNC_BY_USER[user_id] = utc_now()
        return True
    if utc_now() - last >= timedelta(seconds=MEMBERSHIP_SYNC_COOLDOWN_SECONDS):
        LAST_MEMBERSHIP_SYNC_BY_USER[user_id] = utc_now()
        return True
    return False


def sync_membership_for_user_cached(connection, user_row, force=False):
    if not user_row:
        return
    user_id = int(user_row["id"])
    if not should_sync_membership_now(user_id, force=force):
        return
    sync_membership_for_user(connection, user_row)


def sync_membership_for_user(connection, user_row):
    if not STRIPE_SECRET_KEY:
        return

    def priority(item):
        status = item.get("status", "")
        order = {
            "active": 0,
            "trialing": 1,
            "past_due": 2,
            "unpaid": 3,
            "canceled": 4,
            "incomplete": 5,
            "incomplete_expired": 6,
        }
        return (order.get(status, 99), -(item.get("created") or 0))

    customer_candidates = []
    seen_customers = set()

    existing_customer_id = str(user_row["stripe_customer_id"] or "").strip()
    if existing_customer_id:
        seen_customers.add(existing_customer_id)
        customer_candidates.append(existing_customer_id)

    try:
        customers = stripe.Customer.list(email=user_row["email"], limit=20)
        for customer in customers.get("data", []):
            customer_id = str(customer.get("id") or "").strip()
            if not customer_id or customer_id in seen_customers:
                continue
            seen_customers.add(customer_id)
            customer_candidates.append(customer_id)
    except Exception:
        pass

    if not customer_candidates:
        return

    best_subscription = None
    best_customer_id = ""
    best_rank = (99, 0)

    for customer_id in customer_candidates:
        try:
            subscriptions = stripe.Subscription.list(customer=customer_id, status="all", limit=20)
        except Exception:
            continue
        subscription_data = subscriptions.get("data", [])
        if not subscription_data:
            continue
        candidate = sorted(subscription_data, key=priority)[0]
        candidate_rank = priority(candidate)
        if candidate_rank < best_rank:
            best_rank = candidate_rank
            best_subscription = candidate
            best_customer_id = customer_id

    if not best_subscription:
        return

    if best_customer_id and best_customer_id != existing_customer_id:
        connection.execute(
            "UPDATE users SET stripe_customer_id = ?, updated_at = ? WHERE id = ?",
            (best_customer_id, format_timestamp(utc_now()), user_row["id"]),
        )
        connection.commit()

    sync_membership_from_subscription(connection, best_subscription, fallback_user_id=user_row["id"])


def sync_user_membership_safe(connection, user_id):
    if not STRIPE_SECRET_KEY:
        return
    try:
        user_row = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user_row:
            return
        sync_membership_for_user_cached(connection, user_row, force=True)
    except Exception:
        # Admin prehľad musí zostať dostupný aj keď Stripe dočasne zlyhá.
        return


def sync_admin_membership_snapshot(connection, max_users=80):
    if not STRIPE_SECRET_KEY:
        return
    global LAST_ADMIN_SYNC_AT
    if LAST_ADMIN_SYNC_AT and utc_now() - LAST_ADMIN_SYNC_AT < timedelta(seconds=ADMIN_SYNC_COOLDOWN_SECONDS):
        return
    LAST_ADMIN_SYNC_AT = utc_now()
    candidate_rows = connection.execute(
        """
        SELECT DISTINCT users.id
        FROM users
        LEFT JOIN memberships ON memberships.user_id = users.id
        LEFT JOIN checkout_sessions ON checkout_sessions.user_id = users.id
        WHERE
            COALESCE(users.stripe_customer_id, '') != ''
            OR COALESCE(memberships.stripe_subscription_id, '') != ''
            OR COALESCE(memberships.last_checkout_session_id, '') != ''
            OR COALESCE(checkout_sessions.stripe_session_id, '') != ''
        ORDER BY users.updated_at DESC
        LIMIT ?
        """,
        (max(1, int(max_users)),),
    ).fetchall()
    for row in candidate_rows:
        sync_user_membership_safe(connection, int(row["id"]))


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_urls_from_text(text):
    urls = []
    seen = set()
    for match in re.findall(r"https?://[^\s)<]+", str(text or "")):
        cleaned = match.rstrip(".,);]")
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            urls.append(cleaned)
    return urls


def extract_response_source_urls(payload):
    urls = []
    seen = set()

    def add_url(value):
        url = str(value or "").strip()
        if not url or not url.startswith(("http://", "https://")) or url in seen:
            return
        seen.add(url)
        urls.append(url)

    def walk(node):
        if isinstance(node, dict):
            node_type = str(node.get("type") or "").strip().lower()
            if node_type in {"url_citation", "citation", "source"} and node.get("url"):
                add_url(node.get("url"))
            annotations = node.get("annotations")
            if isinstance(annotations, list):
                for item in annotations:
                    walk(item)
            if node.get("url") and ("title" in node or node_type in {"url_citation", "citation", "source"}):
                add_url(node.get("url"))
            for value in node.values():
                if isinstance(value, (dict, list)):
                    walk(value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(payload.get("output") if isinstance(payload, dict) else payload)
    return urls


def estimate_web_confidence(source_urls):
    unique_domains = []
    domain_seen = set()
    trusted_count = 0
    for url in source_urls:
        try:
            host = (urlparse(url).hostname or "").lower().lstrip(".")
        except Exception:
            host = ""
        if not host:
            continue
        host = host[4:] if host.startswith("www.") else host
        if host not in domain_seen:
            domain_seen.add(host)
            unique_domains.append(host)
        if any(host == trusted or host.endswith(f".{trusted}") for trusted in TRUSTED_SOURCE_DOMAINS):
            trusted_count += 1

    source_count = len(unique_domains)
    if source_count == 0:
        return 0

    confidence = 58
    confidence += min(source_count, 4) * 7
    if source_count >= 2:
        confidence += 8
    if source_count >= 3:
        confidence += 4
    confidence += min(trusted_count, 2) * 8
    return min(confidence, 94)


def infer_python_cell_type(value):
    if value is None:
        return "empty"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def table_from_csv(file_bytes):
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSV sa nepodarilo dekódovať.")

    reader = csv.reader(io.StringIO(text), delimiter=detect_csv_delimiter(text))
    rows = [row for row in reader if any(normalize_cell(cell) for cell in row)]
    return table_from_rows(rows)


def detect_csv_delimiter(text):
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        if dialect and dialect.delimiter:
            return dialect.delimiter
    except csv.Error:
        pass

    lines = [line for line in sample.splitlines() if line.strip()][:20]
    if not lines:
        return ","

    candidates = [";", ",", "\t", "|"]
    best = ","
    best_score = -1
    for delimiter in candidates:
        counts = [line.count(delimiter) for line in lines]
        score = sum(counts) + sum(1 for count in counts if count > 0) * 2
        if score > best_score:
            best_score = score
            best = delimiter
    return best


def table_from_xlsx(file_bytes):
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    row_types = []
    for row in sheet.iter_rows():
        values = []
        types = []
        for cell in row:
            values.append(normalize_cell(cell.value))
            if cell.is_date:
                types.append("datetime")
            else:
                types.append(infer_python_cell_type(cell.value))
        rows.append(values)
        row_types.append(types)
    return table_from_rows(rows, row_types)


def table_from_xls(file_bytes):
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    rows = []
    row_types = []
    for row_index in range(sheet.nrows):
        current_row = []
        current_types = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            current_row.append(normalize_cell(cell.value))
            if cell.ctype == xlrd.XL_CELL_DATE:
                current_types.append("datetime")
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                current_types.append("number")
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                current_types.append("empty")
            else:
                current_types.append("text")
        rows.append(current_row)
        row_types.append(current_types)
    return table_from_rows(rows, row_types)


def table_from_rows(rows, row_types=None):
    if not rows:
        raise ValueError("Súbor je prázdny.")

    header_row = rows[0]
    headers = []
    for index, value in enumerate(header_row):
        label = normalize_cell(value) or f"Stĺpec {index + 1}"
        headers.append(label)

    structured_rows = []
    structured_types = []
    for row in rows[1:]:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        structured_rows.append([normalize_cell(padded[index]) for index in range(len(headers))])
    if row_types:
        for row in row_types[1:]:
            padded_types = list(row) + ["empty"] * max(0, len(headers) - len(row))
            structured_types.append([padded_types[index] for index in range(len(headers))])
    else:
        structured_types = [["text" if cell else "empty" for cell in row] for row in structured_rows]

    filtered_rows = []
    filtered_types = []
    for index, row in enumerate(structured_rows):
        if any(cell for cell in row):
            filtered_rows.append(row)
            filtered_types.append(structured_types[index])

    return {
        "headers": headers,
        "rows": filtered_rows,
        "types": filtered_types,
    }


def parse_multipart_form(headers, body):
    content_type = headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type.lower():
        raise ValueError("Požiadavka musí byť multipart/form-data.")

    raw_message = (
        f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8")
        + body
    )
    message = BytesParser(policy=policy.default).parsebytes(raw_message)

    fields = {}
    for part in message.iter_parts():
        if part.get_content_disposition() != "form-data":
            continue

        name = part.get_param("name", header="content-disposition")
        if not name:
            continue

        entry = {
            "filename": part.get_filename(),
            "content": part.get_payload(decode=True) or b"",
            "content_type": part.get_content_type(),
        }
        fields.setdefault(name, []).append(entry)

    return fields


def reject_large_upload(handler, file_size, limit_bytes, label):
    if file_size <= limit_bytes:
        return False
    handler.write_json(
        {
            "error": f"{label} je príliš veľký. Maximum je {format_mb(limit_bytes)}."
        },
        status=HTTPStatus.REQUEST_ENTITY_TOO_LARGE,
    )
    return True


class AppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def maybe_redirect_to_canonical_host(self):
        if not APP_BASE_URL:
            return False
        try:
            configured = APP_BASE_URL if "://" in APP_BASE_URL else f"https://{APP_BASE_URL}"
            canonical = urlparse(configured)
            canonical_host = normalize_host_name(canonical.netloc or "")
            if not canonical_host:
                return False
            request_host = (
                self.headers.get("X-Forwarded-Host")
                or self.headers.get("Host")
                or ""
            ).strip().lower()
            normalized_request_host = normalize_host_name(request_host)
            if not normalized_request_host or normalized_request_host == canonical_host:
                return False
            # Prevent redirect loops between apex and www variants.
            # Canonical redirect is applied only for Render subdomains.
            if not normalized_request_host.endswith(".onrender.com"):
                return False
            destination = f"{APP_BASE_URL.rstrip('/')}{self.path}"
            self.send_response(HTTPStatus.MOVED_PERMANENTLY)
            self.send_header("Location", destination)
            self.end_headers()
            return True
        except Exception:
            return False

    def do_GET(self):
        if self.maybe_redirect_to_canonical_host():
            return
        if self.path == "/ads.txt":
            self.handle_ads_txt()
            return
        if self.path.startswith("/api/me"):
            self.handle_me()
            return
        if self.path.startswith("/api/account"):
            self.handle_account()
            return
        if self.path.startswith("/api/contact-status"):
            self.handle_contact_status()
            return
        if self.path.startswith("/api/assistant-mobile-upload-status"):
            self.handle_assistant_mobile_upload_status()
            return
        if self.path.startswith("/api/assistant"):
            self.handle_assistant()
            return
        if self.path.startswith("/api/generated-asset"):
            self.handle_generated_asset_download()
            return
        if self.path.startswith("/api/admin/user-detail"):
            self.handle_admin_user_detail()
            return
        if self.path.startswith("/api/admin/overview"):
            self.handle_admin_overview()
            return
        if self.path.startswith("/api/refresh-membership"):
            self.handle_refresh_membership()
            return
        super().do_GET()

    def do_POST(self):
        if self.path == "/api/register":
            self.handle_register()
            return
        if self.path == "/api/login":
            self.handle_login()
            return
        if self.path == "/api/request-password-reset":
            self.handle_request_password_reset()
            return
        if self.path == "/api/reset-password":
            self.handle_reset_password()
            return
        if self.path == "/api/logout":
            self.handle_logout()
            return
        if self.path == "/api/account/update":
            self.handle_account_update()
            return
        if self.path == "/api/account/change-password":
            self.handle_account_change_password()
            return
        if self.path == "/api/account/cancel-subscription":
            self.handle_cancel_subscription()
            return
        if self.path == "/api/contact":
            self.handle_contact()
            return
        if self.path == "/api/assistant/thread":
            self.handle_assistant_thread()
            return
        if self.path == "/api/assistant/chat":
            self.handle_assistant_chat()
            return
        if self.path == "/api/assistant-mobile-upload-session":
            self.handle_assistant_mobile_upload_session()
            return
        if self.path == "/api/assistant-mobile-upload":
            self.handle_assistant_mobile_upload()
            return
        if self.path == "/api/assistant/profile":
            self.handle_assistant_profile()
            return
        if self.path == "/api/assistant/tasks":
            self.handle_assistant_tasks()
            return
        if self.path == "/api/admin/assistant-review":
            self.handle_admin_assistant_review()
            return
        if self.path == "/api/admin/assistant-memory-reset":
            self.handle_admin_assistant_memory_reset()
            return
        if self.path == "/api/create-checkout-session":
            self.handle_create_checkout_session()
            return
        if self.path == "/api/admin/membership":
            self.handle_admin_membership_update()
            return
        if self.path == "/api/admin/role":
            self.handle_admin_role_update()
            return
        if self.path == "/api/stripe-webhook":
            self.handle_stripe_webhook()
            return
        if self.path == "/api/process":
            self.handle_process()
            return
        if self.path == "/api/parse":
            self.handle_parse()
            return
        if self.path == "/api/compress-file":
            self.handle_compress_file()
            return
        if self.path == "/api/export-xlsx":
            self.handle_export_xlsx()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Endpoint neexistuje.")

    def read_request_body(self):
        try:
            content_length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            raise ValueError("Neplatná Content-Length hlavička.")
        if content_length <= 0:
            raise ValueError("Požiadavka neobsahuje žiadne dáta.")
        if content_length > MAX_REQUEST_BODY_BYTES:
            raise ValueError(f"Požiadavka je príliš veľká. Maximum je {format_mb(MAX_REQUEST_BODY_BYTES)}.")
        return self.rfile.read(content_length)

    def write_json(self, payload, status=HTTPStatus.OK, cookie_headers=None):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        if cookie_headers:
            for value in cookie_headers:
                self.send_header("Set-Cookie", value)
        self.end_headers()
        self.wfile.write(body)

    def handle_generated_asset_download(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.send_error(HTTPStatus.UNAUTHORIZED, "Najprv sa prihláste.")
                return
            params = parse_qs(urlparse(self.path).query)
            token = str((params.get("token") or [""])[0]).strip()
            if not token:
                self.send_error(HTTPStatus.BAD_REQUEST, "Chýba token súboru.")
                return
            row = get_generated_asset(connection, user["id"], token)
            if not row:
                self.send_error(HTTPStatus.NOT_FOUND, "Súbor už nie je dostupný.")
                return
            body = base64.b64decode(row["payload_base64"])
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", row["content_type"] or "application/octet-stream")
            self.send_header("Content-Disposition", f'attachment; filename="{row["file_name"]}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        finally:
            connection.close()

    def handle_ads_txt(self):
        ads_path = os.path.join(BASE_DIR, "public", "ads.txt")
        if not os.path.exists(ads_path):
            self.send_error(HTTPStatus.NOT_FOUND, "ads.txt neexistuje.")
            return
        with open(ads_path, "rb") as handle:
            body = handle.read()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def is_secure_request(self):
        if APP_BASE_URL.startswith("https://"):
            return True
        forwarded_proto = self.headers.get("X-Forwarded-Proto", "")
        return forwarded_proto.lower() == "https"

    def session_cookie_domain_attr(self):
        request_host = normalize_host_name(
            self.headers.get("X-Forwarded-Host")
            or self.headers.get("Host")
            or ""
        )
        if request_host.endswith("unifyo.online"):
            return " Domain=.unifyo.online;"
        return ""

    def session_cookie_header(self, token, expires_at):
        secure_flag = " Secure;" if self.is_secure_request() else ""
        domain_attr = self.session_cookie_domain_attr()
        return (
            f"{SESSION_COOKIE_NAME}={token}; Path=/;{domain_attr} HttpOnly; SameSite=Lax;{secure_flag} Max-Age={SESSION_TTL_SECONDS}; "
            f"Expires={expires_at.astimezone(timezone.utc).strftime('%a, %d %b %Y %H:%M:%S GMT')}"
        )

    def expired_session_cookie_header(self):
        secure_flag = " Secure;" if self.is_secure_request() else ""
        domain_attr = self.session_cookie_domain_attr()
        return (
            f"{SESSION_COOKIE_NAME}=; Path=/;{domain_attr} HttpOnly; SameSite=Lax;{secure_flag} Max-Age=0; "
            "Expires=Thu, 01 Jan 1970 00:00:00 GMT"
        )

    def handle_me(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"authenticated": False, "user": None})
                return
            user = ensure_admin_role(connection, user)
            try:
                sync_membership_for_user_cached(connection, user, force=False)
                user = connection.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
            except Exception:
                pass
            self.write_json(user_payload(connection, user, self.headers))
        finally:
            connection.close()

    def require_authenticated_user(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return None
            return dict(user)
        finally:
            connection.close()

    def handle_account(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            user = ensure_admin_role(connection, user)
            try:
                sync_membership_for_user_cached(connection, user, force=False)
                user = connection.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
            except Exception:
                pass
            self.write_json(
                {
                    **user_payload(connection, user, self.headers),
                    "activity": get_recent_activity(connection, user["id"], limit=25),
                    "subscription": get_subscription_snapshot(connection, user["id"]),
                }
            )
        finally:
            connection.close()

    def handle_contact_status(self):
        self.write_json(
            {
                "configured": bool(SMTP_HOST and SMTP_USER and SMTP_PASSWORD and SMTP_FROM_EMAIL),
                "reply_to": SMTP_FROM_EMAIL or "",
            }
        )

    def handle_contact(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        contact_name = str(payload.get("contact_name") or "").strip()
        contact_email = normalize_email(payload.get("contact_email"))
        subject = str(payload.get("subject") or "").strip()
        message_text = str(payload.get("message") or "").strip()

        if len(contact_name) < 2:
            self.write_json({"error": "Zadajte meno alebo kontakt."}, status=HTTPStatus.BAD_REQUEST)
            return
        if not contact_email or "@" not in contact_email:
            self.write_json({"error": "Zadajte platný e-mail."}, status=HTTPStatus.BAD_REQUEST)
            return
        if len(subject) < 3:
            self.write_json({"error": "Zadajte predmet správy."}, status=HTTPStatus.BAD_REQUEST)
            return
        if len(message_text) < 10:
            self.write_json({"error": "Správa je príliš krátka."}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            send_contact_email(contact_name, contact_email, subject, message_text)
            log_activity(
                connection,
                "contact_form_sent",
                "Používateľ odoslal kontaktný formulár.",
                user["id"] if user else None,
                contact_email=contact_email,
                subject=subject,
                ip=get_request_ip(self),
                agent=get_request_agent(self),
            )
            self.write_json({"ok": True})
        except RuntimeError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        except smtplib.SMTPException:
            self.write_json({"error": "Kontaktnú správu sa nepodarilo odoslať."}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        finally:
            connection.close()

    def handle_assistant_mobile_upload_session(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": "Na použitie AI asistenta je potrebné aktívne členstvo."}, status=HTTPStatus.PAYMENT_REQUIRED)
                return
            upload_session = create_assistant_upload_session(connection, user["id"])
            base_url = get_request_base_url(self)
            upload_url = f"{base_url}/ai-mobile-upload.html?token={upload_session['token']}"
            self.write_json(
                {
                    "ok": True,
                    "token": upload_session["token"],
                    "upload_url": upload_url,
                    "expires_at": upload_session["expires_at"],
                }
            )
        finally:
            connection.close()

    def handle_assistant_mobile_upload_status(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            params = parse_qs(urlparse(self.path).query)
            token = str((params.get("token") or [""])[0]).strip()
            if not token:
                self.write_json({"error": "Chýba token uploadu."}, status=HTTPStatus.BAD_REQUEST)
                return
            row = get_assistant_upload_session(connection, token)
            if not row:
                self.write_json({"error": "Upload session neexistuje alebo vypršala."}, status=HTTPStatus.NOT_FOUND)
                return
            if row["user_id"] != user["id"]:
                self.write_json({"error": "K tejto upload session nemáte prístup."}, status=HTTPStatus.FORBIDDEN)
                return
            self.write_json({"ok": True, **build_assistant_upload_payload(row)})
        finally:
            connection.close()

    def handle_assistant_mobile_upload(self):
        connection = get_db()
        try:
            form = parse_multipart_form(self.headers, self.read_request_body())
            token = ""
            if form.get("token"):
                token = (form["token"][0].get("content") or b"").decode("utf-8", errors="ignore").strip()
            if not token:
                self.write_json({"error": "Chýba token uploadu."}, status=HTTPStatus.BAD_REQUEST)
                return
            upload_session = get_assistant_upload_session(connection, token)
            if not upload_session:
                self.write_json({"error": "Upload session neexistuje alebo vypršala."}, status=HTTPStatus.NOT_FOUND)
                return
            attachments = parse_assistant_attachments(form)
            if not attachments:
                raise ValueError("Chýba obrázok na nahratie.")
            attachment = attachments[0]
            save_assistant_upload_to_session(connection, token, attachment)
            self.write_json({"ok": True, **build_assistant_upload_payload(get_assistant_upload_session(connection, token))})
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        finally:
            connection.close()

    def handle_assistant(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": "Na použitie AI asistenta je potrebné aktívne členstvo."}, status=HTTPStatus.PAYMENT_REQUIRED)
                return

            params = parse_qs(urlparse(self.path).query)
            requested_thread_id = parse_positive_int((params.get("thread_id") or ["0"])[0], 0)
            active_thread = resolve_assistant_thread(connection, user["id"], requested_thread_id)
            threads = get_assistant_threads(connection, user["id"], limit=50)

            self.write_json(
                {
                    "profile": get_assistant_profile(connection, user["id"]),
                    "threads": threads,
                    "active_thread_id": active_thread["id"] if active_thread else 0,
                    "messages": get_assistant_messages(connection, user["id"], active_thread["id"], limit=80) if active_thread else [],
                }
            )
        finally:
            connection.close()

    def handle_admin_overview(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            user = ensure_admin_role(connection, user)
            if user["role"] != "admin":
                self.write_json({"error": "Nemáte prístup do administrácie."}, status=HTTPStatus.FORBIDDEN)
                return

            try:
                sync_admin_membership_snapshot(connection, max_users=80)
            except Exception:
                # Admin prehľad nesmie byť blokovaný pri dočasnej Stripe chybe.
                pass

            users = []
            removed_accounts = []
            activity = []
            stats = {
                "total_users": 0,
                "active_memberships": 0,
                "admin_users": 0,
                "recent_registrations": 0,
                "ai_threads": 0,
                "ai_messages": 0,
                "ai_active_users": 0,
                "total_logs": 0,
                "recent_logs": 0,
                "recent_unique_ips": 0,
                "recent_admin_actions": 0,
                "deleted_accounts": 0,
                "deactivated_accounts": 0,
            }

            try:
                rows = connection.execute(
                    """
                    SELECT users.*
                    FROM users
                    ORDER BY users.created_at DESC
                    """
                ).fetchall()
                for row in rows:
                    membership_row = get_membership(connection, row["id"])
                    membership_state = get_membership_display_state(membership_row)
                    users.append(
                        {
                            "id": row["id"],
                            "name": row["name"],
                            "email": row["email"],
                            "role": row["role"],
                            "created_at": row["created_at"],
                            "membership_status": membership_state["status"],
                            "membership_valid_until": membership_state["valid_until"],
                        }
                    )
            except Exception:
                users = []

            try:
                removed_rows = connection.execute(
                    """
                    SELECT event_type, event_label, event_meta, created_at
                    FROM activity_logs
                    WHERE event_type IN ('admin_account_deleted', 'admin_membership_deactivated')
                    ORDER BY created_at DESC
                    LIMIT 24
                    """
                ).fetchall()
                for row in removed_rows:
                    meta = safe_json_loads(row["event_meta"], {})
                    removed_accounts.append(
                        {
                            "event_type": row["event_type"],
                            "label": row["event_label"],
                            "email": str(meta.get("target_user_email") or "").strip(),
                            "ip": str(meta.get("ip") or "").strip(),
                            "membership_status": str(meta.get("membership_status") or "").strip(),
                            "membership_valid_until": str(meta.get("membership_valid_until") or "").strip(),
                            "stripe_subscription_id": str(meta.get("stripe_subscription_id") or "").strip(),
                            "last_order_number": str(meta.get("last_order_number") or "").strip(),
                            "created_at": row["created_at"],
                        }
                    )
            except Exception:
                removed_accounts = []

            try:
                activity = get_recent_activity(connection, limit=50)
            except Exception:
                activity = []

            try:
                threshold = format_timestamp(utc_now() - timedelta(days=30))
                stats = {
                    "total_users": len(users),
                    "active_memberships": sum(1 for item in users if item["membership_status"] == "active"),
                    "admin_users": sum(1 for item in users if item["role"] == "admin"),
                    "recent_registrations": connection.execute(
                        "SELECT COUNT(*) AS total FROM users WHERE created_at >= ?",
                        (threshold,),
                    ).fetchone()["total"],
                    "ai_threads": connection.execute("SELECT COUNT(*) AS total FROM assistant_threads").fetchone()["total"],
                    "ai_messages": connection.execute("SELECT COUNT(*) AS total FROM assistant_messages").fetchone()["total"],
                    "ai_active_users": connection.execute(
                        "SELECT COUNT(DISTINCT user_id) AS total FROM assistant_messages WHERE created_at >= ?",
                        (threshold,),
                    ).fetchone()["total"],
                }
                stats.update(get_activity_system_stats(connection))
            except Exception:
                pass

            self.write_json(
                {
                    "users": users,
                    "activity": activity,
                    "removed_accounts": removed_accounts,
                    "stats": stats,
                }
            )
        finally:
            connection.close()

    def handle_admin_user_detail(self):
        connection = get_db()
        try:
            admin_user = get_session_user(connection, self.headers)
            if not admin_user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            admin_user = ensure_admin_role(connection, admin_user)
            if admin_user["role"] != "admin":
                self.write_json({"error": "Nemáte prístup do administrácie."}, status=HTTPStatus.FORBIDDEN)
                return

            params = parse_qs(urlparse(self.path).query)
            user_id = parse_positive_int((params.get("user_id") or ["0"])[0], 0)
            if not user_id:
                self.write_json({"error": "Chýba ID používateľa."}, status=HTTPStatus.BAD_REQUEST)
                return

            # Force per-user sync before detail render so admin sees current state.
            sync_user_membership_safe(connection, user_id)

            row = connection.execute("SELECT users.* FROM users WHERE users.id = ?", (user_id,)).fetchone()
            if not row:
                self.write_json({"error": "Používateľ neexistuje."}, status=HTTPStatus.NOT_FOUND)
                return

            membership_row = get_membership(connection, row["id"])
            membership_state = get_membership_display_state(membership_row)
            registration_consent = get_latest_registration_consent(connection, row["id"]) or {}
            checkout_consent = get_latest_checkout_consent(connection, row["id"]) or {}
            params = parse_qs(urlparse(self.path).query)
            self.write_json(
                {
                    "user": {
                        "id": row["id"],
                        "name": row["name"],
                        "email": row["email"],
                        "role": row["role"],
                        "created_at": row["created_at"],
                        "membership_status": membership_state["status"],
                        "membership_valid_until": membership_state["valid_until"],
                    },
                    "activity": get_recent_activity(connection, user_id=row["id"], limit=25),
                    "assistant_stats": get_assistant_usage_stats(connection, row["id"]),
                    "subscription": get_subscription_snapshot(connection, row["id"]),
                    "registration_consent": {
                        "created_at": registration_consent.get("created_at") or "",
                        "ip_address": registration_consent.get("ip_address") or "",
                        "user_agent": registration_consent.get("user_agent") or "",
                        "marketing_consent": bool(registration_consent.get("marketing_consent")),
                        "consent_text": registration_consent.get("consent_text") or "",
                        "legal_version": registration_consent.get("legal_version") or "",
                    },
                    "checkout_consent": {
                        "created_at": checkout_consent.get("created_at") or "",
                        "ip_address": checkout_consent.get("ip_address") or "",
                        "user_agent": checkout_consent.get("user_agent") or "",
                        "consent_text": checkout_consent.get("consent_text") or "",
                        "consent_version": checkout_consent.get("consent_version") or "",
                        "price_label": checkout_consent.get("price_label") or "",
                        "subscription_label": checkout_consent.get("subscription_label") or "",
                        "renewal_label": checkout_consent.get("renewal_label") or "",
                        "no_refund_label": checkout_consent.get("no_refund_label") or "",
                    },
                }
            )
        finally:
            connection.close()

    def handle_refresh_membership(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            sync_membership_for_user_cached(connection, user, force=True)
            user = connection.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
            self.write_json(user_payload(connection, user, self.headers))
        except stripe.error.StripeError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_GATEWAY)
        finally:
            connection.close()

    def handle_register(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        name = str(payload.get("name") or "").strip()
        email = normalize_email(payload.get("email"))
        password = str(payload.get("password") or "")
        accept_terms = bool(payload.get("accept_terms"))
        accept_privacy = bool(payload.get("accept_privacy"))
        marketing_consent = bool(payload.get("marketing_consent"))
        if len(name) < 2:
            self.write_json({"error": "Zadajte meno alebo názov používateľa."}, status=HTTPStatus.BAD_REQUEST)
            return
        if not email or "@" not in email:
            self.write_json({"error": "Zadajte platný e-mail."}, status=HTTPStatus.BAD_REQUEST)
            return
        if len(password) < 8:
            self.write_json({"error": "Heslo musí mať aspoň 8 znakov."}, status=HTTPStatus.BAD_REQUEST)
            return
        if not accept_terms or not accept_privacy:
            self.write_json(
                {"error": "Pre vytvorenie účtu musíte potvrdiť obchodné podmienky a GDPR."},
                status=HTTPStatus.BAD_REQUEST,
            )
            return

        connection = get_db()
        try:
            existing = connection.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
            if existing:
                self.write_json({"error": "Účet s týmto e-mailom už existuje."}, status=HTTPStatus.CONFLICT)
                return

            salt, password_hash = create_password_hash(password)
            now = format_timestamp(utc_now())
            role = "admin" if is_admin_email(email) else "user"
            cursor = connection.execute(
                """
                INSERT INTO users (name, email, password_hash, password_salt, role, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (name, email, password_hash, salt, role, now, now),
            )
            user_id = cursor.lastrowid
            record_registration_consent(
                connection,
                user_id,
                email,
                get_request_ip(self),
                get_request_agent(self),
                marketing_consent,
            )
            token, expires_at = create_session(connection, user_id)
            log_activity(
                connection,
                "register",
                "Používateľ si vytvoril účet.",
                user_id,
                email=email,
                role=role,
                ip=get_request_ip(self),
                agent=get_request_agent(self),
                marketing_consent=marketing_consent,
                legal_version=LEGAL_VERSION,
            )
            connection.commit()
            user = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
            try:
                send_registration_welcome_email(user)
            except Exception:
                pass
            self.write_json(
                user_payload(connection, user, self.headers),
                status=HTTPStatus.CREATED,
                cookie_headers=[self.session_cookie_header(token, expires_at)],
            )
        finally:
            connection.close()

    def handle_login(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        email = normalize_email(payload.get("email"))
        password = str(payload.get("password") or "")
        connection = get_db()
        try:
            user = connection.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
            if not user or not verify_password(password, user["password_salt"], user["password_hash"]):
                self.write_json({"error": "Neplatný e-mail alebo heslo."}, status=HTTPStatus.UNAUTHORIZED)
                return
            user = ensure_admin_role(connection, user)

            old_token = get_cookie_token(self.headers)
            clear_session(connection, old_token)
            token, expires_at = create_session(connection, user["id"])
            log_activity(connection, "login", "Používateľ sa prihlásil.", user["id"], email=email, ip=get_request_ip(self), agent=get_request_agent(self))
            self.write_json(
                user_payload(connection, user, self.headers),
                cookie_headers=[self.session_cookie_header(token, expires_at)],
            )
        finally:
            connection.close()

    def handle_request_password_reset(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        email = normalize_email(payload.get("email"))
        if not email:
            self.write_json({"error": "Zadajte e-mail."}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            user = connection.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
            if not user:
                self.write_json({"ok": True})
                return

            raw_token = secrets.token_urlsafe(32)
            token_hash = hash_session_token(raw_token)
            now = utc_now()
            expires_at = now + timedelta(minutes=30)
            connection.execute("DELETE FROM password_reset_tokens WHERE user_id = ?", (user["id"],))
            connection.execute(
                """
                INSERT INTO password_reset_tokens (user_id, token_hash, expires_at, created_at, used_at)
                VALUES (?, ?, ?, ?, NULL)
                """,
                (user["id"], token_hash, format_timestamp(expires_at), format_timestamp(now)),
            )
            connection.commit()
            send_reset_email(user["email"], user["name"], raw_token)
            log_activity(connection, "password_reset_requested", "Používateľ si vyžiadal obnovu hesla.", user["id"])
            self.write_json({"ok": True})
        except RuntimeError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        except smtplib.SMTPAuthenticationError:
            self.write_json(
                {"error": "SMTP prihlásenie zlyhalo. Skontrolujte SMTP_USER a SMTP_PASSWORD."},
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        except smtplib.SMTPException:
            self.write_json(
                {"error": "Odoslanie reset e-mailu zlyhalo na SMTP serveri."},
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        except (TimeoutError, socket.timeout):
            self.write_json(
                {"error": "SMTP server neodpovedá. Skontrolujte SMTP_HOST a SMTP_PORT."},
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        except Exception as exc:
            self.write_json({"error": f"Reset e-mail sa nepodarilo odoslať: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        finally:
            connection.close()

    def handle_reset_password(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        token = str(payload.get("token") or "").strip()
        password = str(payload.get("password") or "")
        if len(password) < 8:
            self.write_json({"error": "Nové heslo musí mať aspoň 8 znakov."}, status=HTTPStatus.BAD_REQUEST)
            return
        if not token:
            self.write_json({"error": "Chýba token obnovy hesla."}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            token_row = connection.execute(
                "SELECT * FROM password_reset_tokens WHERE token_hash = ?",
                (hash_session_token(token),),
            ).fetchone()
            if not token_row:
                self.write_json({"error": "Reset link je neplatný."}, status=HTTPStatus.BAD_REQUEST)
                return
            if token_row["used_at"]:
                self.write_json({"error": "Reset link už bol použitý."}, status=HTTPStatus.BAD_REQUEST)
                return
            if parse_timestamp(token_row["expires_at"]) <= utc_now():
                self.write_json({"error": "Reset link už vypršal."}, status=HTTPStatus.BAD_REQUEST)
                return

            salt, password_hash = create_password_hash(password)
            now = format_timestamp(utc_now())
            connection.execute(
                "UPDATE users SET password_hash = ?, password_salt = ?, updated_at = ? WHERE id = ?",
                (password_hash, salt, now, token_row["user_id"]),
            )
            connection.execute(
                "UPDATE password_reset_tokens SET used_at = ? WHERE id = ?",
                (now, token_row["id"]),
            )
            connection.execute("DELETE FROM sessions WHERE user_id = ?", (token_row["user_id"],))
            connection.commit()
            log_activity(connection, "password_reset_completed", "Používateľ si obnovil heslo.", token_row["user_id"])
            self.write_json({"ok": True})
        finally:
            connection.close()

    def handle_account_update(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return

            name = str(payload.get("name") or "").strip()
            email = normalize_email(payload.get("email"))
            if len(name) < 2:
                self.write_json({"error": "Zadajte meno alebo názov používateľa."}, status=HTTPStatus.BAD_REQUEST)
                return
            if not email or "@" not in email:
                self.write_json({"error": "Zadajte platný e-mail."}, status=HTTPStatus.BAD_REQUEST)
                return

            existing = connection.execute(
                "SELECT id FROM users WHERE email = ? AND id != ?",
                (email, user["id"]),
            ).fetchone()
            if existing:
                self.write_json({"error": "Tento e-mail už používa iný účet."}, status=HTTPStatus.CONFLICT)
                return

            role = user["role"]
            connection.execute(
                "UPDATE users SET name = ?, email = ?, role = ?, updated_at = ? WHERE id = ?",
                (name, email, role, format_timestamp(utc_now()), user["id"]),
            )
            connection.commit()
            log_activity(connection, "account_updated", "Používateľ upravil svoj účet.", user["id"], email=email)
            updated = connection.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
            self.write_json({**user_payload(connection, updated), "activity": get_recent_activity(connection, updated["id"], limit=25)})
        finally:
            connection.close()

    def handle_account_change_password(self):
        try:
            payload = require_json(self)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        current_password = str(payload.get("current_password") or "")
        new_password = str(payload.get("new_password") or "")
        if len(new_password) < 8:
            self.write_json({"error": "Nové heslo musí mať aspoň 8 znakov."}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            if not verify_password(current_password, user["password_salt"], user["password_hash"]):
                self.write_json({"error": "Aktuálne heslo nie je správne."}, status=HTTPStatus.BAD_REQUEST)
                return

            salt, password_hash = create_password_hash(new_password)
            connection.execute(
                "UPDATE users SET password_hash = ?, password_salt = ?, updated_at = ? WHERE id = ?",
                (password_hash, salt, format_timestamp(utc_now()), user["id"]),
            )
            connection.commit()
            log_activity(connection, "password_changed", "Používateľ zmenil heslo v účte.", user["id"])
            self.write_json({"ok": True})
        finally:
            connection.close()

    def handle_assistant_profile(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": "Na použitie AI asistenta je potrebné aktívne členstvo."}, status=HTTPStatus.PAYMENT_REQUIRED)
                return

            log_activity(connection, "assistant_profile_checked", "Používateľ otvoril nastavenie AI profilu.", user["id"])
            self.write_json({"ok": True, "profile": get_assistant_profile(connection, user["id"])})
        finally:
            connection.close()

    def handle_assistant_thread(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": "Na použitie AI asistenta je potrebné aktívne členstvo."}, status=HTTPStatus.PAYMENT_REQUIRED)
                return

            payload = require_json(self)
            action = str(payload.get("action") or "create").strip().lower()
            language = "en" if str(payload.get("lang") or "").strip().lower() == "en" else "sk"

            if action == "rename":
                thread_id = parse_positive_int(payload.get("thread_id"), 0)
                title = build_thread_title(payload.get("title") or ("New chat" if language == "en" else "Nový chat"), language=language)
                active_thread = resolve_assistant_thread(connection, user["id"], thread_id)
                if not active_thread:
                    self.write_json({"error": "Chat sa nenašiel."}, status=HTTPStatus.NOT_FOUND)
                    return
                touch_assistant_thread(connection, active_thread["id"], title=title)
                connection.commit()
                active_thread = resolve_assistant_thread(connection, user["id"], active_thread["id"])
                log_activity(connection, "assistant_thread_renamed", "Používateľ premenoval AI chat.", user["id"], thread_id=active_thread["id"], title=title)
                self.write_json(
                    {
                        "ok": True,
                        "thread": {
                            "id": active_thread["id"],
                            "title": active_thread["title"] or "Nový chat",
                            "created_at": active_thread["created_at"],
                            "updated_at": active_thread["updated_at"],
                        },
                        "threads": get_assistant_threads(connection, user["id"], limit=50),
                        "messages": get_assistant_messages(connection, user["id"], active_thread["id"], limit=80),
                    }
                )
                return

            if action == "delete":
                thread_id = parse_positive_int(payload.get("thread_id"), 0)
                deleted_thread = delete_assistant_thread(connection, user["id"], thread_id)
                if not deleted_thread:
                    self.write_json({"error": "Chat sa nenašiel."}, status=HTTPStatus.NOT_FOUND)
                    return
                active_thread = get_latest_assistant_thread(connection, user["id"])
                log_activity(connection, "assistant_thread_deleted", "Používateľ vymazal AI chat.", user["id"], thread_id=thread_id)
                self.write_json(
                    {
                        "ok": True,
                        "threads": get_assistant_threads(connection, user["id"], limit=50),
                        "active_thread_id": active_thread["id"] if active_thread else 0,
                        "messages": get_assistant_messages(connection, user["id"], active_thread["id"], limit=80) if active_thread else [],
                    }
                )
                return

            title = build_thread_title(payload.get("title") or ("New chat" if language == "en" else "Nový chat"), language=language)
            thread_id = create_assistant_thread(connection, user["id"], title)
            active_thread = resolve_assistant_thread(connection, user["id"], thread_id)
            log_activity(connection, "assistant_thread_created", "Používateľ založil nový AI chat.", user["id"], thread_id=thread_id)
            self.write_json(
                {
                    "ok": True,
                    "thread": {
                        "id": active_thread["id"],
                        "title": active_thread["title"] or "Nový chat",
                        "created_at": active_thread["created_at"],
                        "updated_at": active_thread["updated_at"],
                    },
                    "threads": get_assistant_threads(connection, user["id"], limit=50),
                    "messages": [],
                }
            )
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        finally:
            connection.close()

    def handle_assistant_chat(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": "Na použitie AI asistenta je potrebné aktívne členstvo."}, status=HTTPStatus.PAYMENT_REQUIRED)
                return

            content_type = self.headers.get("Content-Type", "")
            attachments = []
            if "multipart/form-data" in content_type.lower():
                form = parse_multipart_form(self.headers, self.read_request_body())
                message = ""
                language = "sk"
                if form.get("message"):
                    message = (form["message"][0].get("content") or b"").decode("utf-8", errors="ignore").strip()
                thread_id = 0
                if form.get("thread_id"):
                    thread_id = parse_positive_int((form["thread_id"][0].get("content") or b"").decode("utf-8", errors="ignore").strip(), 0)
                if form.get("lang"):
                    language = "en" if ((form["lang"][0].get("content") or b"").decode("utf-8", errors="ignore").strip().lower() == "en") else "sk"
                attachments = parse_assistant_attachments(form)
            else:
                payload = require_json(self)
                message = str(payload.get("message") or "").strip()
                thread_id = parse_positive_int(payload.get("thread_id"), 0)
                language = "en" if str(payload.get("lang") or "").strip().lower() == "en" else "sk"
            if len(message) < 2 and not attachments:
                self.write_json({"error": "Napíšte správu pre AI asistenta."}, status=HTTPStatus.BAD_REQUEST)
                return

            active_thread = resolve_assistant_thread(connection, user["id"], thread_id)
            if not active_thread:
                new_thread_id = create_assistant_thread(connection, user["id"], build_thread_title(message, language=language))
                active_thread = resolve_assistant_thread(connection, user["id"], new_thread_id)
            profile = get_assistant_profile(connection, user["id"])
            profile["brief"] = build_assistant_brief(get_assistant_tasks(connection, user["id"], include_done=True))
            history_messages = get_assistant_messages(connection, user["id"], active_thread["id"], limit=80)
            user_message_content = message
            user_meta = {}
            if attachments:
                base_user_message = message or "Vyhodnoť prosím priložené obrázky."
                user_message_content = base_user_message
                user_meta = {"attachments": []}
                for attachment in attachments:
                    user_meta["attachments"].append(
                        {
                            "attachment_name": attachment["filename"],
                            "attachment_type": attachment["content_type"],
                            "attachment_preview": attachment["data_url"],
                        }
                    )
                primary_attachment = user_meta["attachments"][0]
                user_meta["attachment_name"] = primary_attachment["attachment_name"]
                user_meta["attachment_type"] = primary_attachment["attachment_type"]
                user_meta["attachment_preview"] = primary_attachment["attachment_preview"]
            save_assistant_message(connection, user["id"], active_thread["id"], "user", user_message_content, meta=user_meta, language=language)
            reply, reply_meta = call_openai_assistant(user, profile, history_messages, message, attachments=attachments, language=language)
            parsed_table = parse_markdown_table(reply)
            if parsed_table:
                reply_meta["table_data"] = parsed_table
            if should_generate_pdf_asset(message, reply):
                pdf_title = derive_thread_title(
                    history_messages + [{"role": "user", "content": message}, {"role": "assistant", "content": reply}],
                    language=language,
                    existing_title=active_thread["title"],
                )
                pdf_bytes = build_pdf_bytes(pdf_title, strip_markdown_table_lines(reply) or reply, table_data=parsed_table)
                pdf_name = sanitize_download_filename(pdf_title, "unifyo_ai_export")
                pdf_token = store_generated_asset(
                    connection,
                    user["id"],
                    f"{pdf_name}.pdf",
                    "application/pdf",
                    pdf_bytes,
                    kind="pdf",
                )
                reply_meta["generated_asset_name"] = f"{pdf_name}.pdf"
                reply_meta["generated_asset_url"] = f"/api/generated-asset?token={pdf_token}"
                reply_meta["generated_asset_kind"] = "pdf"
            save_assistant_message(connection, user["id"], active_thread["id"], "assistant", reply, meta=reply_meta, language=language)
            updated_messages = get_assistant_messages(connection, user["id"], active_thread["id"], limit=80)
            next_title = derive_thread_title(updated_messages, language=language, existing_title=active_thread["title"])
            touch_assistant_thread(connection, active_thread["id"], title=next_title)
            connection.commit()
            refresh_assistant_profile_memory(connection, user["id"])
            log_activity(
                connection,
                "assistant_chat_message",
                "Používateľ komunikoval s AI asistentom.",
                user["id"],
                thread_id=active_thread["id"],
                used_image=bool(attachments),
                attachment_name=attachments[0]["filename"] if attachments else "",
                attachment_count=len(attachments),
                ip=get_request_ip(self),
                agent=get_request_agent(self),
            )
            self.write_json(
                {
                    "ok": True,
                    "threads": get_assistant_threads(connection, user["id"], limit=50),
                    "active_thread_id": active_thread["id"],
                    "messages": get_assistant_messages(connection, user["id"], active_thread["id"], limit=80),
                    "profile": profile,
                }
            )
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        except RuntimeError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_GATEWAY)
        finally:
            connection.close()

    def handle_assistant_tasks(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": "Na použitie AI asistenta je potrebné aktívne členstvo."}, status=HTTPStatus.PAYMENT_REQUIRED)
                return

            payload = require_json(self)
            action = str(payload.get("action") or "create").strip().lower()

            if action == "create":
                title = str(payload.get("title") or "").strip()
                client_name = str(payload.get("client_name") or "").strip()
                due_date = str(payload.get("due_date") or "").strip()
                priority = str(payload.get("priority") or "medium").strip().lower()
                channel = str(payload.get("channel") or "followup").strip().lower()
                details = str(payload.get("details") or "").strip()

                if len(title) < 3:
                    self.write_json({"error": "Zadajte názov úlohy."}, status=HTTPStatus.BAD_REQUEST)
                    return
                if priority not in {"high", "medium", "low"}:
                    priority = "medium"
                if channel not in {"followup", "email", "call", "meeting"}:
                    channel = "followup"

                now_value = format_timestamp(utc_now())
                connection.execute(
                    """
                    INSERT INTO assistant_tasks (
                        user_id, title, client_name, due_date, priority, channel, details, status, completed_at, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 'open', '', ?, ?)
                    """,
                    (user["id"], title, client_name, due_date, priority, channel, details, now_value, now_value),
                )
                connection.commit()
                log_activity(connection, "assistant_task_created", "Používateľ pridal úlohu do AI asistenta.", user["id"], title=title, client_name=client_name, priority=priority)
                self.write_json({"ok": True})
                return

            task_id = parse_positive_int(payload.get("task_id"), 0)
            if not task_id:
                self.write_json({"error": "Chýba ID úlohy."}, status=HTTPStatus.BAD_REQUEST)
                return

            task_row = connection.execute(
                "SELECT * FROM assistant_tasks WHERE id = ? AND user_id = ?",
                (task_id, user["id"]),
            ).fetchone()
            if not task_row:
                self.write_json({"error": "Úloha neexistuje."}, status=HTTPStatus.NOT_FOUND)
                return

            if action == "complete":
                connection.execute(
                    "UPDATE assistant_tasks SET status = 'done', completed_at = ?, updated_at = ? WHERE id = ?",
                    (format_timestamp(utc_now()), format_timestamp(utc_now()), task_id),
                )
                connection.commit()
                log_activity(connection, "assistant_task_completed", "Používateľ označil úlohu ako hotovú.", user["id"], task_id=task_id, title=task_row["title"])
                self.write_json({"ok": True})
                return

            if action == "reopen":
                connection.execute(
                    "UPDATE assistant_tasks SET status = 'open', completed_at = '', updated_at = ? WHERE id = ?",
                    (format_timestamp(utc_now()), task_id),
                )
                connection.commit()
                log_activity(connection, "assistant_task_reopened", "Používateľ znova otvoril úlohu.", user["id"], task_id=task_id, title=task_row["title"])
                self.write_json({"ok": True})
                return

            if action == "delete":
                connection.execute("DELETE FROM assistant_tasks WHERE id = ?", (task_id,))
                connection.commit()
                log_activity(connection, "assistant_task_deleted", "Používateľ vymazal úlohu z AI asistenta.", user["id"], task_id=task_id, title=task_row["title"])
                self.write_json({"ok": True})
                return

            self.write_json({"error": "Neplatná akcia pre AI asistenta."}, status=HTTPStatus.BAD_REQUEST)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        finally:
            connection.close()

    def handle_logout(self):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if user:
                log_activity(connection, "logout", "Používateľ sa odhlásil.", user["id"], ip=get_request_ip(self), agent=get_request_agent(self))
            clear_session(connection, get_cookie_token(self.headers))
            self.write_json(
                {"ok": True},
                cookie_headers=[self.expired_session_cookie_header()],
            )
        finally:
            connection.close()

    def handle_create_checkout_session(self):
        price_id = resolve_stripe_price_id()
        if not STRIPE_SECRET_KEY or not price_id:
            self.write_json(
                {"error": "Stripe nie je nakonfigurovaný. Nastav STRIPE_SECRET_KEY a STRIPE_PRICE_ID."},
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
            return
        if not re.fullmatch(r"price_[A-Za-z0-9]+", price_id):
            self.write_json(
                {
                    "error": (
                        "Stripe cena nie je nastavená správne. "
                        "Premenná STRIPE_PRICE_ID musí byť vo formáte price_... "
                        "(nie payment link URL)."
                    )
                },
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )
            return

        connection = get_db()
        try:
            payload = require_json(self)
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste alebo zaregistrujte."}, status=HTTPStatus.UNAUTHORIZED)
                return

            membership = get_membership(connection, user["id"])
            if is_membership_active(membership):
                self.write_json({"error": "Členstvo je už aktívne."}, status=HTTPStatus.BAD_REQUEST)
                return
            if not bool(payload.get("checkout_consent_accepted")):
                self.write_json(
                    {"error": "Pred pokračovaním musíte potvrdiť súhlas so začatím poskytovania služby."},
                    status=HTTPStatus.BAD_REQUEST,
                )
                return

            base_url = get_request_base_url(self)
            customer_id = get_or_create_customer(connection, user)
            internal_order_number = generate_internal_order_number()
            internal_subscription_number = (membership or {}).get("internal_subscription_number") or generate_internal_subscription_number()
            consent_id = record_checkout_consent(connection, user, get_request_ip(self), get_request_agent(self))
            session = stripe.checkout.Session.create(
                mode="subscription",
                customer=customer_id,
                line_items=[{"price": price_id, "quantity": 1}],
                success_url=f"{base_url}/app.html?checkout=success",
                cancel_url=f"{base_url}/app.html?checkout=cancel",
                allow_promotion_codes=True,
                metadata={
                    "user_id": str(user["id"]),
                    "internal_order_number": internal_order_number,
                    "internal_subscription_number": internal_subscription_number,
                    "checkout_consent_version": CHECKOUT_CONSENT_VERSION,
                },
                subscription_data={
                    "metadata": {
                        "user_id": str(user["id"]),
                        "internal_order_number": internal_order_number,
                        "internal_subscription_number": internal_subscription_number,
                    }
                },
                client_reference_id=str(user["id"]),
            )
            connection.execute(
                """
                INSERT OR REPLACE INTO checkout_sessions (user_id, stripe_session_id, created_at, internal_order_number, checkout_consent_id)
                VALUES (?, ?, ?, ?, ?)
                """,
                (user["id"], session["id"], format_timestamp(utc_now()), internal_order_number, consent_id),
            )
            connection.execute(
                """
                INSERT INTO memberships (
                    user_id, status, stripe_customer_id, stripe_subscription_id, current_period_end, created_at, updated_at,
                    stripe_status, cancel_at_period_end, cancelled_at, next_renewal_at,
                    internal_subscription_number, last_order_number, last_checkout_session_id, last_payment_intent_id, last_invoice_id
                ) VALUES (?, 'pending', ?, '', '', ?, ?, 'pending', 0, '', '', ?, ?, ?, '', '')
                ON CONFLICT(user_id) DO UPDATE SET
                    stripe_customer_id = excluded.stripe_customer_id,
                    internal_subscription_number = excluded.internal_subscription_number,
                    last_order_number = excluded.last_order_number,
                    last_checkout_session_id = excluded.last_checkout_session_id,
                    updated_at = excluded.updated_at
                """,
                (
                    user["id"],
                    customer_id,
                    format_timestamp(utc_now()),
                    format_timestamp(utc_now()),
                    internal_subscription_number,
                    internal_order_number,
                    session["id"],
                ),
            )
            record_subscription_event(
                connection,
                user["id"],
                "checkout_started",
                "pending",
                {
                    "internal_order_number": internal_order_number,
                    "internal_subscription_number": internal_subscription_number,
                    "stripe_session_id": session["id"],
                    "consent_version": CHECKOUT_CONSENT_VERSION,
                },
            )
            connection.commit()
            log_activity(
                connection,
                "checkout_created",
                "Používateľ otvoril Stripe checkout pre členstvo.",
                user["id"],
                internal_order_number=internal_order_number,
                internal_subscription_number=internal_subscription_number,
                stripe_session_id=session["id"],
                ip=get_request_ip(self),
                agent=get_request_agent(self),
            )
            self.write_json({"url": session["url"]})
        except stripe.error.StripeError as exc:
            error_text = str(exc)
            if "expected pattern" in error_text.lower():
                error_text = (
                    "Stripe odmietol identifikátor ceny. "
                    "Skontrolujte STRIPE_PRICE_ID (musí byť price_...)."
                )
            self.write_json({"error": error_text}, status=HTTPStatus.BAD_GATEWAY)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        finally:
            connection.close()

    def handle_cancel_subscription(self):
        if not STRIPE_SECRET_KEY:
            self.write_json({"error": "Stripe nie je nakonfigurovaný."}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            return

        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            membership = get_membership(connection, user["id"])
            if not membership or not membership.get("stripe_subscription_id"):
                self.write_json({"error": "Predplatné sa nenašlo."}, status=HTTPStatus.NOT_FOUND)
                return

            subscription = stripe.Subscription.modify(
                membership["stripe_subscription_id"],
                cancel_at_period_end=True,
            )
            sync_membership_from_subscription(connection, subscription, fallback_user_id=user["id"])
            updated_membership = get_membership(connection, user["id"]) or {}
            record_subscription_event(
                connection,
                user["id"],
                "subscription_cancel_requested",
                updated_membership.get("status") or "active",
                {
                    "stripe_subscription_id": updated_membership.get("stripe_subscription_id") or "",
                    "access_until": format_timestamp(updated_membership.get("current_period_end")) if updated_membership.get("current_period_end") else "",
                    "cancel_at_period_end": bool(updated_membership.get("cancel_at_period_end")),
                },
            )
            send_subscription_cancelled_email(
                user,
                {
                    "access_until": format_timestamp(updated_membership.get("current_period_end")) if updated_membership.get("current_period_end") else "",
                    "internal_order_number": updated_membership.get("last_order_number") or "",
                    "internal_subscription_number": updated_membership.get("internal_subscription_number") or "",
                    "stripe_subscription_id": updated_membership.get("stripe_subscription_id") or "",
                },
            )
            log_activity(
                connection,
                "subscription_cancelled",
                "Používateľ zrušil predplatné ku koncu obdobia.",
                user["id"],
                stripe_subscription_id=updated_membership.get("stripe_subscription_id") or "",
                access_until=format_timestamp(updated_membership.get("current_period_end")) if updated_membership.get("current_period_end") else "",
                ip=get_request_ip(self),
                agent=get_request_agent(self),
            )
            connection.commit()
            self.write_json(
                {
                    "ok": True,
                    "message": f"Predplatné bolo úspešne zrušené. Prístup máte aktívny do {format_timestamp(updated_membership.get('current_period_end')) if updated_membership.get('current_period_end') else ''}.",
                    "access_until": format_timestamp(updated_membership.get("current_period_end")) if updated_membership.get("current_period_end") else "",
                }
            )
        except stripe.error.StripeError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_GATEWAY)
        finally:
            connection.close()

    def handle_admin_membership_update(self):
        connection = get_db()
        try:
            admin_user = get_session_user(connection, self.headers)
            if not admin_user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            admin_user = ensure_admin_role(connection, admin_user)
            if admin_user["role"] != "admin":
                self.write_json({"error": "Nemáte prístup do administrácie."}, status=HTTPStatus.FORBIDDEN)
                return
            if not can_manage_admin_tools(admin_user):
                self.write_json({"error": "Len hlavný správca môže meniť členstvá používateľov."}, status=HTTPStatus.FORBIDDEN)
                return

            payload = require_json(self)
            user_id = int(payload.get("user_id") or 0)
            action = str(payload.get("action") or "").strip()
            if not user_id or action not in {"delete_account", "deactivate"}:
                self.write_json({"error": "Neplatná admin akcia."}, status=HTTPStatus.BAD_REQUEST)
                return

            target_user = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
            if not target_user:
                self.write_json({"error": "Používateľ neexistuje."}, status=HTTPStatus.NOT_FOUND)
                return
            membership_before = get_membership(connection, user_id) or {}

            target_email = target_user["email"] or ""
            target_name = target_user["name"] or ""
            warning_text = "Účet bude nenávratne vymazaný. Ak bola platba strhnutá, refund bude spracovaný do 1 týždňa."
            if action == "deactivate":
                now_value = format_timestamp(utc_now())
                connection.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
                connection.execute(
                    """
                    UPDATE memberships
                    SET status = 'inactive',
                        stripe_status = 'inactive',
                        cancel_at_period_end = 1,
                        cancelled_at = COALESCE(cancelled_at, ?),
                        current_period_end = '',
                        next_renewal_at = '',
                        updated_at = ?
                    WHERE user_id = ?
                    """,
                    (now_value, now_value, user_id),
                )
                record_subscription_event(
                    connection,
                    user_id,
                    "admin_deactivated",
                    "inactive",
                    {
                        "admin_email": admin_user["email"],
                        "target_user_email": target_email,
                    },
                )
                log_activity(
                    connection,
                    "admin_membership_deactivated",
                    "Admin deaktivoval používateľský účet a členstvo.",
                    admin_user["id"],
                    target_user_id=user_id,
                    target_user_email=target_email,
                    membership_status=membership_before.get("status") or "",
                    membership_valid_until=format_timestamp(membership_before.get("current_period_end")) if membership_before.get("current_period_end") else "",
                    stripe_subscription_id=membership_before.get("stripe_subscription_id") or "",
                    last_order_number=membership_before.get("last_order_number") or "",
                    ip=get_request_ip(self),
                    agent=get_request_agent(self),
                )
            else:
                connection.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM memberships WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM assistant_tasks WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM assistant_messages WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM assistant_threads WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM assistant_profiles WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM assistant_upload_sessions WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM generated_assets WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM registration_consents WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM checkout_consents WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM checkout_sessions WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM subscription_events WHERE user_id = ?", (user_id,))
                connection.execute("DELETE FROM users WHERE id = ?", (user_id,))
                log_activity(
                    connection,
                    "admin_account_deleted",
                    "Admin vymazal používateľský účet.",
                    admin_user["id"],
                    target_user_id=user_id,
                    target_user_email=target_email,
                    warning=warning_text,
                    membership_status=membership_before.get("status") or "",
                    membership_valid_until=format_timestamp(membership_before.get("current_period_end")) if membership_before.get("current_period_end") else "",
                    stripe_subscription_id=membership_before.get("stripe_subscription_id") or "",
                    last_order_number=membership_before.get("last_order_number") or "",
                    ip=get_request_ip(self),
                    agent=get_request_agent(self),
                )
                try:
                    send_account_deleted_email(target_email, target_name)
                except Exception:
                    pass
            connection.commit()
            self.write_json({"ok": True})
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        finally:
            connection.close()

    def handle_admin_role_update(self):
        connection = get_db()
        try:
            admin_user = get_session_user(connection, self.headers)
            if not admin_user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            admin_user = ensure_admin_role(connection, admin_user)
            if admin_user["role"] != "admin":
                self.write_json({"error": "Nemáte prístup do administrácie."}, status=HTTPStatus.FORBIDDEN)
                return
            if not can_manage_admin_tools(admin_user):
                self.write_json({"error": "Len hlavný správca môže meniť roly používateľov."}, status=HTTPStatus.FORBIDDEN)
                return

            payload = require_json(self)
            user_id = parse_positive_int(payload.get("user_id"), 0)
            role = str(payload.get("role") or "").strip().lower()
            if not user_id or role not in {"admin", "user"}:
                self.write_json({"error": "Neplatná rola alebo používateľ."}, status=HTTPStatus.BAD_REQUEST)
                return

            target_user = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
            if not target_user:
                self.write_json({"error": "Používateľ neexistuje."}, status=HTTPStatus.NOT_FOUND)
                return
            if target_user["id"] == admin_user["id"] and role != "admin":
                self.write_json({"error": "Nemôžete odobrať admin práva sami sebe."}, status=HTTPStatus.BAD_REQUEST)
                return
            if role == "user" and is_admin_email(target_user["email"]):
                self.write_json({"error": "E-mail správcu je rezervovaný pre administrátorský účet."}, status=HTTPStatus.BAD_REQUEST)
                return

            connection.execute(
                "UPDATE users SET role = ?, updated_at = ? WHERE id = ?",
                (role, format_timestamp(utc_now()), user_id),
            )
            connection.commit()
            log_activity(connection, "admin_role_update", f"Admin zmenil rolu používateľa na {role}.", admin_user["id"], target_user_id=user_id, role=role)
            self.write_json({"ok": True})
        finally:
            connection.close()

    def handle_admin_assistant_review(self):
        connection = get_db()
        try:
            admin_user = get_session_user(connection, self.headers)
            if not admin_user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            admin_user = ensure_admin_role(connection, admin_user)
            if admin_user["role"] != "admin":
                self.write_json({"error": "Nemáte prístup do administrácie."}, status=HTTPStatus.FORBIDDEN)
                return

            payload = require_json(self)
            user_id = parse_positive_int(payload.get("user_id"), 0)
            message_id = parse_positive_int(payload.get("message_id"), 0)
            review_status = str(payload.get("review_status") or "").strip().lower()
            if not user_id or not message_id or review_status not in {"approved", "needs_review"}:
                self.write_json({"error": "Neplatná AI review akcia."}, status=HTTPStatus.BAD_REQUEST)
                return

            message_row = connection.execute(
                """
                SELECT *
                FROM assistant_messages
                WHERE id = ? AND user_id = ? AND role = 'assistant'
                """,
                (message_id, user_id),
            ).fetchone()
            if not message_row:
                self.write_json({"error": "AI odpoveď sa nenašla."}, status=HTTPStatus.NOT_FOUND)
                return

            connection.execute(
                """
                UPDATE assistant_messages
                SET review_status = ?, reviewed_by = ?, reviewed_at = ?
                WHERE id = ?
                """,
                (review_status, admin_user["id"], format_timestamp(utc_now()), message_id),
            )
            connection.commit()
            log_activity(
                connection,
                "assistant_review_updated",
                "Admin zmenil stav AI odpovede.",
                admin_user["id"],
                target_user_id=user_id,
                message_id=message_id,
                review_status=review_status,
            )
            self.write_json({"ok": True})
        finally:
            connection.close()

    def handle_admin_assistant_memory_reset(self):
        connection = get_db()
        try:
            admin_user = get_session_user(connection, self.headers)
            if not admin_user:
                self.write_json({"error": "Najprv sa prihláste."}, status=HTTPStatus.UNAUTHORIZED)
                return
            admin_user = ensure_admin_role(connection, admin_user)
            if admin_user["role"] != "admin":
                self.write_json({"error": "Nemáte prístup do administrácie."}, status=HTTPStatus.FORBIDDEN)
                return

            payload = require_json(self)
            user_id = parse_positive_int(payload.get("user_id"), 0)
            if not user_id:
                self.write_json({"error": "Chýba používateľ."}, status=HTTPStatus.BAD_REQUEST)
                return

            target_user = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
            if not target_user:
                self.write_json({"error": "Používateľ neexistuje."}, status=HTTPStatus.NOT_FOUND)
                return

            connection.execute("DELETE FROM assistant_messages WHERE user_id = ?", (user_id,))
            connection.execute("DELETE FROM assistant_threads WHERE user_id = ?", (user_id,))
            connection.execute(
                """
                INSERT INTO assistant_profiles (user_id, focus, notes, updated_at)
                VALUES (?, '', '', ?)
                ON CONFLICT(user_id) DO UPDATE SET
                    focus = '',
                    notes = '',
                    updated_at = excluded.updated_at
                """,
                (user_id, format_timestamp(utc_now())),
            )
            connection.commit()
            log_activity(
                connection,
                "assistant_memory_reset",
                "Admin vymazal AI históriu a pamäť používateľa.",
                admin_user["id"],
                target_user_id=user_id,
            )
            self.write_json({"ok": True})
        finally:
            connection.close()

    def handle_stripe_webhook(self):
        if not STRIPE_WEBHOOK_SECRET:
            self.write_json({"error": "Webhook secret nie je nastavený."}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            return

        payload = self.read_request_body()
        signature = self.headers.get("Stripe-Signature", "")

        try:
            event = stripe.Webhook.construct_event(payload, signature, STRIPE_WEBHOOK_SECRET)
        except ValueError:
            self.write_json({"error": "Neplatný webhook payload."}, status=HTTPStatus.BAD_REQUEST)
            return
        except stripe.error.SignatureVerificationError:
            self.write_json({"error": "Neplatný Stripe podpis."}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            event_type = event["type"]
            data = event["data"]["object"]

            if event_type == "checkout.session.completed":
                user_id = int(data.get("metadata", {}).get("user_id") or data.get("client_reference_id") or 0)
                customer_email = normalize_email(
                    (data.get("customer_details") or {}).get("email")
                    or data.get("customer_email")
                    or ""
                )
                if not user_id and customer_email:
                    matched_user = connection.execute(
                        "SELECT * FROM users WHERE email = ?",
                        (customer_email,),
                    ).fetchone()
                    if matched_user:
                        user_id = int(matched_user["id"])
                subscription_id = data.get("subscription")
                internal_order_number = data.get("metadata", {}).get("internal_order_number") or ""
                internal_subscription_number = data.get("metadata", {}).get("internal_subscription_number") or ""
                payment_intent_id = data.get("payment_intent") or ""
                session_id = data.get("id") or ""
                checkout_row = connection.execute(
                    "SELECT * FROM checkout_sessions WHERE stripe_session_id = ?",
                    (session_id,),
                ).fetchone()
                if subscription_id:
                    subscription = stripe.Subscription.retrieve(subscription_id)
                    sync_membership_from_subscription(connection, subscription, fallback_user_id=user_id or None)
                    user_row = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
                    if user_row:
                        connection.execute(
                            """
                            UPDATE memberships
                            SET last_order_number = ?,
                                last_checkout_session_id = ?,
                                last_payment_intent_id = ?,
                                internal_subscription_number = CASE WHEN internal_subscription_number = '' THEN ? ELSE internal_subscription_number END,
                                last_invoice_id = CASE WHEN ? != '' THEN ? ELSE last_invoice_id END,
                                updated_at = ?
                            WHERE user_id = ?
                            """,
                            (
                                internal_order_number,
                                session_id,
                                payment_intent_id,
                                internal_subscription_number,
                                subscription.get("latest_invoice") or "",
                                subscription.get("latest_invoice") or "",
                                format_timestamp(utc_now()),
                                user_id,
                            ),
                        )
                        record_subscription_event(
                            connection,
                            user_id,
                            "checkout_completed",
                            "active",
                            {
                                "internal_order_number": internal_order_number,
                                "internal_subscription_number": internal_subscription_number,
                                "stripe_session_id": session_id,
                                "stripe_subscription_id": subscription_id,
                                "stripe_payment_intent_id": payment_intent_id,
                            },
                        )
                        updated_membership = get_membership(connection, user_id) or {}
                        send_subscription_activated_email(
                            user_row,
                            {
                                "price_label": "1,99 € / mesiac",
                                "purchased_at": format_timestamp(utc_now()),
                                "next_renewal_at": format_timestamp(updated_membership.get("next_renewal_at")) if updated_membership.get("next_renewal_at") else "",
                                "internal_order_number": internal_order_number,
                                "internal_subscription_number": updated_membership.get("internal_subscription_number") or internal_subscription_number,
                                "stripe_subscription_id": updated_membership.get("stripe_subscription_id") or subscription_id,
                                "stripe_payment_intent_id": payment_intent_id,
                            },
                        )
                if checkout_row:
                    connection.execute(
                        """
                        UPDATE checkout_sessions
                        SET internal_order_number = CASE WHEN internal_order_number = '' THEN ? ELSE internal_order_number END
                        WHERE id = ?
                        """,
                        (internal_order_number, checkout_row["id"]),
                    )

            if event_type in {"customer.subscription.created", "customer.subscription.updated", "customer.subscription.deleted"}:
                sync_membership_from_subscription(connection, data)

            self.write_json({"received": True})
        except Exception as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        finally:
            connection.close()

    def require_active_membership(self, membership_error_message="Na nahranie vlastných súborov je potrebné aktívne členstvo."):
        connection = get_db()
        try:
            user = get_session_user(connection, self.headers)
            if not user:
                self.write_json({"error": "Pre túto akciu sa musíte prihlásiť."}, status=HTTPStatus.UNAUTHORIZED)
                return None

            membership = get_membership(connection, user["id"])
            if not user_has_service_access(user, membership):
                self.write_json({"error": membership_error_message}, status=HTTPStatus.PAYMENT_REQUIRED)
                return None
            return user
        finally:
            connection.close()

    def handle_parse(self):
        try:
            form = parse_multipart_form(self.headers, self.read_request_body())
        except Exception as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        file_items = form.get("file", [])
        if not file_items:
            self.write_json({"error": "Súbor nebol prijatý."}, status=HTTPStatus.BAD_REQUEST)
            return

        file_item = file_items[0]
        file_name = file_item.get("filename") or "upload"
        extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
        file_bytes = file_item.get("content", b"")
        if reject_large_upload(self, len(file_bytes), MAX_CONTACT_FILE_BYTES, "Súbor"):
            return

        try:
            if extension == "csv":
                table = table_from_csv(file_bytes)
            elif extension == "xlsx":
                table = table_from_xlsx(file_bytes)
            elif extension == "xls":
                table = table_from_xls(file_bytes)
            else:
                raise ValueError("Podporované sú len CSV, XLSX a XLS.")
        except Exception as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        self.write_json(table)

    def handle_process(self):
        user = self.require_active_membership()
        if not user:
            return

        try:
            form = parse_multipart_form(self.headers, self.read_request_body())
        except Exception as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        if "files" not in form:
            self.write_json({"error": "Neboli prijaté žiadne súbory."}, status=HTTPStatus.BAD_REQUEST)
            return

        uploads = []
        for file_item in form["files"]:
            file_name = file_item.get("filename") or "upload"
            file_bytes = file_item.get("content", b"")
            if file_bytes:
                if reject_large_upload(self, len(file_bytes), MAX_CONTACT_FILE_BYTES, f"Súbor {file_name}"):
                    return
                uploads.append((file_name, file_bytes))

        if not uploads:
            self.write_json({"error": "Neboli prijaté žiadne súbory."}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            payload = process_uploads(uploads)
        except Exception as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        connection = get_db()
        try:
            log_activity(connection, "upload_processed", "Používateľ spracoval databázu kontaktov.", user["id"], files=len(uploads))
        finally:
            connection.close()
        self.write_json(payload)

    def handle_compress_file(self):
        user = self.require_active_membership("Na použitie kompresie súborov je potrebné aktívne členstvo.")
        if not user:
            return

        try:
            form = parse_multipart_form(self.headers, self.read_request_body())

            file_items = form.get("file", [])
            if not file_items:
                self.write_json({"error": "Súbor nebol prijatý."}, status=HTTPStatus.BAD_REQUEST)
                return

            target_items = form.get("target_mb", [])
            target_mb = ""
            if target_items:
                target_mb = (target_items[0].get("content", b"") or b"").decode("utf-8", "ignore").strip()

            file_item = file_items[0]
            file_name = file_item.get("filename") or "upload"
            file_bytes = file_item.get("content", b"")
            if reject_large_upload(self, len(file_bytes), MAX_COMPRESS_FILE_BYTES, "Súbor"):
                return

            result = compress_upload(file_name, file_bytes, target_mb)

            connection = get_db()
            try:
                log_activity(
                    connection,
                    "file_compressed",
                    "Používateľ zmenšil súbor.",
                    user["id"],
                    file=file_name,
                    original_bytes=result.original_bytes,
                    compressed_bytes=result.compressed_bytes,
                    target_bytes=result.target_bytes,
                    status=result.status,
                )
            finally:
                connection.close()

            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", result.content_type)
            self.send_header("Content-Disposition", f'attachment; filename="{result.file_name}"')
            self.send_header("Content-Length", str(len(result.data)))
            self.send_header("X-Compression-Original-Bytes", str(result.original_bytes))
            self.send_header("X-Compression-Compressed-Bytes", str(result.compressed_bytes))
            self.send_header("X-Compression-Target-Bytes", str(result.target_bytes))
            self.send_header("X-Compression-Status", result.status)
            self.send_header("X-Compression-Reached-Target", "1" if result.reached_target else "0")
            self.end_headers()
            self.wfile.write(result.data)
        except ValueError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        except RuntimeError as exc:
            self.write_json({"error": str(exc)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
        except Exception as exc:
            self.write_json({"error": f"Kompresia zlyhala: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_export_xlsx(self):
        user = self.require_active_membership()
        if not user:
            return

        try:
            payload = require_json(self)
        except ValueError:
            self.write_json({"error": "Neplatné dáta pre export."}, status=HTTPStatus.BAD_REQUEST)
            return

        rows = payload.get("rows", [])
        filename = payload.get("filename", "merged_contacts.xlsx")
        if not rows:
            self.write_json({"error": "Nie sú dáta na export."}, status=HTTPStatus.BAD_REQUEST)
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Kontakty"

        sheet.append(OUTPUT_HEADERS)
        for row in rows:
            sheet.append([row.get(header, "") for header in OUTPUT_HEADERS])

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)

        buffer = io.BytesIO()
        workbook.save(buffer)
        body = buffer.getvalue()

        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)
        connection = get_db()
        try:
            log_activity(connection, "xlsx_export", "Používateľ exportoval XLSX.", user["id"], filename=filename)
        finally:
            connection.close()


def main():
    init_db()
    server = HTTPServer((HOST, PORT), AppHandler)
    print(f"Serving Unifyo on http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
